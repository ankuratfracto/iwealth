"""Excel export helpers for statements workbooks.

Normalizes parsed statement DataFrames, applies light heuristics for
column ordering and number coercion, discovers period labels, and writes a
single styled workbook (one sheet per canonical document type). Also
provides routing fallbacks and small utilities used by the pipeline.
"""

from __future__ import annotations

from typing import List, Dict, Any
from pathlib import Path
import io, json, time, logging, os, re, sys

from iwe_core.config import CFG
from iwe_core.debug_utils import (
    vprint,
    valdbg_enabled,
    dprint,
    debug_enabled,
    debug_flag_from_cfg,
)
from iwe_core.grouping import normalize_doc_type
from iwe_core.pdf_ops import get_page_count_from_bytes
from iwe_core.json_ops import (
    extract_rows as _extract_rows,
    extract_period_maps_from_payload as _extract_period_maps_from_payload,
    scan_group_jsons_for_periods as _scan_group_jsons_for_periods,
)
from iwe_core import analytics as _analytics
from iwe_core.mid_pass import filter_pages_via_mid_pass
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from concurrent.futures import ThreadPoolExecutor, as_completed

logger = logging.getLogger(__name__)

# Shared cache for period labels used by Excel header renaming
PERIOD_LABELS_BY_DOC: dict[str, dict] = globals().get('PERIOD_LABELS_BY_DOC', {}) or {}

# --- Module defaults and light-weight helpers to keep pyflakes clean ---
# Validation feature flags and settings (pulled from CFG; with sensible defaults)
VALIDATION_SUM_ENABLE = bool(((CFG.get("validation", {}) or {}).get("checks", {}) or {}).get("balance_sheet", {}).get("sum_subitems", {}).get("enable", True))
VALIDATION_SUM_TOL_PCT = float(((CFG.get("validation", {}) or {}).get("checks", {}) or {}).get("balance_sheet", {}).get("sum_subitems", {}).get("tolerance_pct", 0.001))
VALIDATION_SUM_ABS_MIN = float(((CFG.get("validation", {}) or {}).get("checks", {}) or {}).get("balance_sheet", {}).get("sum_subitems", {}).get("abs_min", 1.0))

# Fine-grained validation toggles (with sensible defaults)
_VAL_CFG_BS = (((CFG.get("validation", {}) or {}).get("checks", {}) or {}).get("balance_sheet", {}) or {})
VAL_DECLARED_COMPONENTS = bool(_VAL_CFG_BS.get("declared_components", True))
VAL_CHILDREN_WITHOUT_COMPONENTS = bool(_VAL_CFG_BS.get("children_without_components", True))
VAL_SECTION_CHECKS = bool(_VAL_CFG_BS.get("section_checks", True))
VAL_SECTION_EQUALITY = bool(_VAL_CFG_BS.get("section_equality", True))
VAL_BLOCK_FALLBACK = bool(_VAL_CFG_BS.get("contiguous_block_fallback", True))
VAL_COMPOSED_AND_GRAND = bool(_VAL_CFG_BS.get("composed_and_grand", True))
INCLUDE_VALIDATION_SHEET = bool(((CFG.get("export", {}) or {}).get("statements_workbook", {}) or {}).get("include_validation_sheet", True))
VALIDATION_SHEET_NAME = str(((CFG.get("export", {}) or {}).get("statements_workbook", {}) or {}).get("validation_sheet_name", "Validation"))
VALIDATION_PROFILES = (CFG.get("validation", {}) or {}).get("profiles", {}) or {}

# ---- Small common helpers (deduplicated) ----------------------------------
def _compile(regex_list):
    import re as _re
    out = []
    for p in (regex_list or []):
        try:
            out.append(_re.compile(str(p), _re.I))
        except Exception:
            pass
    return out


def _find_rows_matching(labels_series, patterns):
    hits = []
    try:
        for idx, val in labels_series.items():
            s = str(val or "")
            for pat in patterns:
                if pat.search(s):
                    hits.append(idx)
                    break
    except Exception:
        pass
    return hits


def _first_val_from_rows(df, rows_idx, col):
    for i in rows_idx or []:
        v = _coerce_number_like(df.at[i, col])
        if v is not None:
            return float(v)
    return None


def _parse_components_value(val):
    if val is None:
        return []
    if isinstance(val, list):
        return [str(x).strip() for x in val if str(x).strip()]
    s = str(val).strip()
    if not s:
        return []
    try:
        import json as _json
        arr = _json.loads(s)
        if isinstance(arr, list):
            return [str(x).strip() for x in arr if str(x).strip()]
    except Exception:
        pass
    out = []
    for tok in re.split(r"[,|;]", s):
        tok = tok.strip()
        tok = re.sub(r"^[\[\]\s\"\']+|[\[\]\s\"\']+$", "", tok)
        if tok:
            out.append(tok)
    return out


def _section_label_from_df(df, idx: int, default: str = "Declared") -> str:
    """Prefer section_id, else section/Section/sectionId; fallback to default label."""
    for key in ("section_id", "section", "Section", "sectionId"):
        if key in df.columns:
            try:
                val = str(df.at[idx, key]).strip()
                if val and val.lower() != "nan":
                    return val
            except Exception:
                pass
    return default

# Excel mapping/template defaults
MAPPINGS: dict[str, str] = {}
TEMPLATE_PATH: str | None = None
SHEET_NAME: str | None = None

# OCR routing defaults
EXTRA_ACCURACY_FIRST = "true"
EXTRA_ACCURACY_SECOND = "true"
SECOND_PARSER_APP_ID = ""
SECOND_MODEL_ID = ""

# debug helpers imported from iwe_core.debug_utils

def _labels_only(period_by_doc: dict | None) -> dict:
    """Return labels-only mapping for period ids; tolerant of None/input shapes."""
    out: dict[str, dict] = {}
    for dt, mapping in (period_by_doc or {}).items():
        try:
            out[str(dt)] = {str(k).lower(): (v if isinstance(v, str) else str((v or {}).get("label", ""))) for k, v in (mapping or {}).items()}
        except Exception:
            out[str(dt)] = {}
    return out

def _pick_period_labels_for_sheet(sheet_name: str, labels_by_doc: dict | None, global_cache: dict | None = None) -> dict | None:
    """Find period labels for this sheet.

    Order of preference:
      1) Exact match (case-insensitive) in local labels_by_doc
      2) Exact match in global_cache (from earlier runs)
      3) Fallback: best available doc in (local → global) by number of labels
         (useful when only one doc provided labels; assumes common periods across statements)
    """
    try:
        name_l = str(sheet_name).strip().lower()
        # Search local then global
        sources = []
        if isinstance(labels_by_doc, dict):
            sources.append(labels_by_doc)
        if isinstance(global_cache, dict):
            sources.append(global_cache)
        # 1) Exact match search
        for src in sources:
            if sheet_name in src:
                return src[sheet_name]
            for k, v in src.items():
                try:
                    if str(k).strip().lower() == name_l:
                        return v
                except Exception:
                    continue
        # 3) Best-available fallback by label count
        for src in sources:
            best = None
            best_n = 0
            for k, v in src.items():
                try:
                    n = len(v or {})
                    if n > best_n:
                        best = v
                        best_n = n
                except Exception:
                    continue
            if best:
                return best
    except Exception:
        pass
    return None
# Optional reorders/sanitizers (no-op fallbacks)
def sanitize_statement_df(doc_type: str, df):
    return df

def reorder_dataframe_sections_first(df):
    return df

# OCR / grouping helpers with fallbacks
try:
    from iwe_core.ocr_client import call_fracto, call_fracto_parallel
except Exception:
    def call_fracto(pdf_bytes, file_name, **kwargs):
        return {"file": file_name, "data": {}}
    def call_fracto_parallel(pdf_bytes, file_name, **kwargs):
        return [call_fracto(pdf_bytes, file_name, **kwargs)]

try:
    from iwe_core.grouping import build_groups
except Exception:
    def build_groups(selected_pages, classification, pdf_bytes):
        return {}

def expand_selected_pages(selected_pages: list[int], total_pages: int, radius: int = 1) -> list[int]:
    """Expand selected page numbers by a small neighbour radius (clamped)."""
    s = set()
    for p in selected_pages or []:
        for d in range(-radius, radius + 1):
            q = p + d
            if 1 <= q <= max(1, int(total_pages)):
                s.add(q)
    return sorted(s)

def _resolve_routing(doc_type: str) -> tuple[str, str, str]:
    """Fallback routing resolver for excel_ops context.

    Uses third-pass defaults and logs the choice for visibility.
    """
    d = (CFG.get("passes", {}).get("third", {}).get("defaults", {}) or {})
    parser = d.get("parser_app", "")
    model = d.get("model", "tv6")
    extra = str(d.get("extra_accuracy", True)).lower()
    try:
        logger.info("[routing] excel_ops: using third_defaults for %s → parser=%s model=%s extra=%s", doc_type, parser, model, extra)
    except Exception:
        pass
    return (parser, model, extra)


# json debug helpers imported from iwe_core.debug_utils


def _coerce_number_like(x):
    """
    Try to parse a value that "looks numeric" into a float.
    - Returns None for blank/NA-ish
    - Returns float when possible, preserving sign hints in parentheses
    - Otherwise returns None (caller may keep original value)
    """
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "" or s.lower() in {"na","n/a","nil","none","nan","-","–","—"}:
        return None
    # Handle (123) as -123
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    # Strip commas and common currency symbols
    s = s.replace(",", "").replace("₹", "").replace("$", "").replace("€", "").replace("£", "")
    # Drop any trailing footnote marks or non-numeric chars
    try:
        import re as _re
        s = _re.sub(r"[^0-9.\-]", "", s)
    except Exception:
        pass
    try:
        v = float(s)
        return -v if neg else v
    except Exception:
        return None

def _normalize_df_for_excel(doc_type: str, df: Any) -> Any:
    """
    Prepare a DataFrame for Excel:
      • sanitize_statement_df (merge notes, tidy totals)
      • coerce number-like columns to numeric
      • ensure a 'Particulars' column exists and is first
    """
    if df is None or df.empty:
        return df
    df = sanitize_statement_df(doc_type, df)

    # Reorder columns: sr_no, Particulars, then c1..cN, then the rest
    try:
        import re
        cols = list(df.columns)
        sr = next((c for c in cols if str(c).strip().lower() == "sr_no"), None)
        part_aliases = {"particulars","particular","description","line item","line_item","account head","head"}
        part_col = next((c for c in cols if str(c).strip().lower() in part_aliases), None)
        if part_col and part_col != "Particulars":
            df = df.rename(columns={part_col: "Particulars"})
            part_col = "Particulars"
        if not part_col and "Particulars" in df.columns:
            part_col = "Particulars"
        # collect c-columns
        c_cols = sorted(
            [c for c in df.columns if re.fullmatch(r"[cC]\\d+", str(c))],
            key=lambda x: int(re.findall(r"\\d+", str(x))[0])
        )
        ordered = []
        if sr:
            ordered.append(sr)
        if part_col and part_col not in ordered:
            ordered.append(part_col)
        ordered.extend([c for c in c_cols if c not in ordered])
        ordered.extend([c for c in df.columns if c not in ordered])
        df = df.loc[:, ordered]
    except Exception:
        pass

    # Coerce numbers only in period columns c1..cN (avoid meta like id/components/parent_id)
    try:
        import re as _re
        period_cols = [c for c in df.columns if _re.fullmatch(r"(?i)c\d+", str(c))]
    except Exception:
        period_cols = [c for c in df.columns if str(c).lower().startswith("c")]
    for c in period_cols:
        coerced = df[c].apply(_coerce_number_like)
        if sum(v is not None for v in coerced) >= max(1, int(0.5 * len(df))):
            df[c] = coerced
    return df

# row extraction delegated to json_ops.extract_rows

def _write_statements_workbook(pdf_path: str, stem: str, combined_sheets: dict[str, Any], routing_used: dict[str, dict] | None = None, periods_by_doc: dict[str, dict] | None = None) -> str:
    """
    Write a single Excel workbook with:
      • Fixed sheet order from config (or canonical labels)
      • Styled headers (colors from config)
      • Autosized columns with wrapping
      • Optional 'Routing Summary' sheet
    Returns the file path.
    """
    import pandas as pd
    from datetime import datetime
    try:
        import openpyxl as _oxl
        _oxl_ver = getattr(_oxl, "__version__", "unknown")
    except Exception:
        _oxl_ver = "unavailable"

    global PERIOD_LABELS_BY_DOC

    use_period_labels_cfg = CFG.get("export", {}).get("statements_workbook", {}).get("use_period_labels", True)
    xlsx_name_tmpl = CFG.get("export", {}).get("filenames", {}).get("statements_xlsx", "{stem}_statements.xlsx")
    out_path = Path(pdf_path).expanduser().resolve().with_name(xlsx_name_tmpl.format(stem=stem))
    print(f"[Excel] ENTER _write_statements_workbook → out={out_path}", flush=True)
    try:
        print(
            f"[Excel] Env: ts={datetime.now().isoformat(timespec='seconds')}, pid={os.getpid()}, cwd={os.getcwd()}\n"
            f"        Versions: python={sys.version.split()[0]}, pandas={pd.__version__}, openpyxl={_oxl_ver}",
            flush=True,
        )
    except Exception:
        pass

    sheet_order = CFG.get("export", {}).get("statements_workbook", {}).get("sheet_order") \
        or CFG.get("labels", {}).get("canonical", []) \
        or sorted(combined_sheets.keys())

    style_cfg = CFG.get("export", {}).get("statements_workbook", {}).get("style", {}) or {}
    header_fill_hex     = str(style_cfg.get("header_fill", "305496")).strip()
    header_font_color   = str(style_cfg.get("header_font_color", "FFFFFF")).strip()
    freeze_panes        = str(CFG.get("export", {}).get("statements_workbook", {}).get("freeze_panes", "A2"))

    # Discover/receive period labels
    print(f"[Excel] use_period_labels_cfg={use_period_labels_cfg}", flush=True)

    period_by_doc = periods_by_doc or {}
    period_labels_by_doc = _labels_only(period_by_doc)

    # If nothing was passed in-memory, poll the disk a few times before giving up.
    if not period_labels_by_doc:
        for attempt in range(1, 8):
            _by_doc, _labels = _scan_group_jsons_for_periods(pdf_path, stem)
            if _labels:
                period_by_doc = _by_doc or {}
                period_labels_by_doc = _labels or {}
                print(f"[Excel] Period labels discovered via disk scan on attempt {attempt}", flush=True)
                break
            else:
                print(f"[Excel] No period labels on disk yet (attempt {attempt}) — will retry shortly...", flush=True)
                time.sleep(1)

    # Final fallback: inspect combined statements JSON from this/prior run
    if not period_labels_by_doc:
        try:
            json_name_tmpl = (CFG.get("export", {}) or {}).get("combined_json", {}).get("filename", "{stem}_statements.json")
            combined_json_path = Path(pdf_path).expanduser().resolve().with_name(json_name_tmpl.format(stem=stem))
            if combined_json_path.exists():
                with open(combined_json_path, "r", encoding="utf-8") as fh:
                    combined_obj = json.load(fh) or {}
                docs = combined_obj.get("documents") or {}
                for dt, entry in docs.items():
                    labels = {}
                    for cid, meta in (entry.get("periods") or {}).items():
                        labels[str(cid).lower()] = (meta or {}).get("label", "")
                    if labels:
                        period_labels_by_doc.setdefault(dt, {}).update(labels)
                print("[Excel] Fallback from combined JSON → period label docs:",
                    list(period_labels_by_doc.keys()), flush=True)
        except Exception as e:
            print("[Excel] Combined JSON fallback failed:", e, flush=True)

    try:
        print("[Excel] Periods (local keys):", list((period_labels_by_doc or {}).keys()), flush=True)
    except Exception:
        pass
    try:
        print("[Excel] Periods (global cache keys):", list((PERIOD_LABELS_BY_DOC or {}).keys()), flush=True)
    except Exception:
        pass

    if period_labels_by_doc:
        for _k, _v in period_labels_by_doc.items():
            PERIOD_LABELS_BY_DOC.setdefault(_k, {}).update(_v)

    debug_dump = {
        "use_period_labels": use_period_labels_cfg,
        "periods_by_doc_labels": _labels_only(period_by_doc),
        "sheets": {}
    }
    print(f"[Excel] sheet-order: {sheet_order}")

    # Quick overview of input sheets
    try:
        for _sn, _df in (combined_sheets or {}).items():
            try:
                _shape = tuple(getattr(_df, "shape", (0, 0)))
                _cols = list(getattr(_df, "columns", []))
                _dtypes = [str(getattr(_df, "dtypes", []).__getitem__(c)) for c in _cols] if hasattr(_df, "dtypes") else []
                _mem = int(_df.memory_usage(index=True, deep=True).sum()) if hasattr(_df, "memory_usage") else -1
                print(f"[Excel] IN[{_sn}] rows={_shape[0]} cols={_shape[1]} mem={_mem}B cols={_cols}", flush=True)
                if _dtypes:
                    print(f"[Excel] IN[{_sn}] dtypes={dict(zip(_cols, _dtypes))}", flush=True)
            except Exception as _e0:
                print(f"[Excel] IN[{_sn}] overview failed: {_e0}", flush=True)
    except Exception:
        pass

    t0 = time.perf_counter()

    def _sanitize_text(s: Any) -> Any:
        try:
            if s is None:
                return None
            if isinstance(s, (int, float)):
                return s
            t = str(s)
            # Remove XML-invalid control characters
            import re as _re
            t = _re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", " ", t)
            return t
        except Exception:
            return s

    def _sanitize_df_for_excel(df):
        try:
            import pandas as _pd
            if df is None or getattr(df, 'empty', True):
                return df
            obj_cols = [c for c in df.columns if str(getattr(df.dtypes, '__getitem__')(c)).startswith('object')]
            if obj_cols:
                df[obj_cols] = df[obj_cols].applymap(_sanitize_text)
        except Exception:
            pass
        return df

    used_sheet_names: set[str] = set()

    def _unique_sheet_name(raw: str) -> str:
        import re as _re
        name = str(raw or 'Sheet')
        # Remove illegal characters
        name = _re.sub(r"[:\\/*?\[\]]", " ", name).strip()
        name = name[:31] or 'Sheet'
        base = name
        idx = 2
        while name in used_sheet_names:
            suffix = f" ({idx})"
            name = (base[: max(0, 31 - len(suffix))] + suffix)[:31]
            idx += 1
        used_sheet_names.add(name)
        return name
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        try:
            print(f"[Excel] Writer created: engine=openpyxl, book_type={type(writer.book).__name__}", flush=True)
        except Exception:
            pass

        # Track whether any real statement/routing/periods sheet was written
        wrote_any_real_sheet = False

        validation_rows: list[list] = []

        # Pre-write one sheet early to guarantee at least one visible sheet exists
        first_written = False
        first_name: str | None = None
        try:
            for cand in sheet_order:
                df0 = combined_sheets.get(cand)
                if df0 is None:
                    continue
                first_name = cand
                try:
                    df_pre = _normalize_df_for_excel(cand, df0.copy())
                except Exception:
                    df_pre = df0
                try:
                    df_pre = _sanitize_df_for_excel(df_pre)
                except Exception:
                    pass
                safe0 = _unique_sheet_name(cand)
                df_pre.to_excel(writer, sheet_name=safe0, index=False)
                wrote_any_real_sheet = True
                first_written = True
                break
        except Exception:
            pass

        def _pick_company_type_from_routing(routing: dict | None) -> str:
            try:
                for _dt, meta in (routing or {}).items():
                    ct = (meta or {}).get("company_type")
                    if ct:
                        return str(ct).strip().lower()
            except Exception:
                pass
            return "corporate"

        # use module-level _compile and _find_rows_matching

        def _coerce_series_num(s):
            return s.apply(_coerce_number_like)

        def _compute_sum_checks_for_bs(sheet_label: str, df_in: "pd.DataFrame") -> None:
            if not VALIDATION_SUM_ENABLE or df_in is None or getattr(df_in, "empty", True):
                return
            import pandas as _pd
            df = df_in.copy()
            # High-level entry log for BS validation
            try:
                vprint({
                    "phase": "BS.validate.enter",
                    "sheet": sheet_label,
                    "shape": list(getattr(df, "shape", [])),
                    "columns": list(map(str, getattr(df, "columns", []))),
                    "has_cols": {
                        "id": "id" in df.columns,
                        "parent_id": "parent_id" in df.columns,
                        "calc_refs": ("calculation_references" in df.columns) or ("components" in df.columns),
                        "row_type": "row_type" in df.columns,
                        "section": "section" in df.columns,
                    }
                })
            except Exception:
                pass
            # Identify columns
            part_col = next((c for c in df.columns if str(c).strip().lower() in {"particulars","particular","description","line item","line_item"}), None)
            if part_col is None:
                return
            # Only validate numeric period columns (c1, c2, ...). Avoid meta columns like id/parent_id/components.
            import re as _re
            num_cols = [c for c in df.columns if _re.fullmatch(r"(?i)c\d+", str(c))]
            if not num_cols:
                # Fallback: any non-meta column with at least one numeric value
                meta_cols = {part_col, "sr_no", "id", "row_type", "parent_id", "components", "calculation_references", "section"}
                num_cols = [c for c in df.columns if str(c) not in meta_cols]
            try:
                # Show numeric counts per candidate column before pruning
                _num_counts = {}
                for c in list(num_cols):
                    try:
                        coerced = df[c].apply(_coerce_number_like)
                        _num_counts[str(c)] = int(sum(v is not None for v in coerced))
                    except Exception:
                        _num_counts[str(c)] = -1
                vprint({
                    "phase": "BS.numeric_columns.candidates",
                    "sheet": sheet_label,
                    "counts": _num_counts,
                })
            except Exception:
                pass
            # Quick debug: show numeric columns picked
            try:
                print(f"[Validation] BS numeric columns for '{sheet_label}': {num_cols}", flush=True)
            except Exception:
                pass
            # helper to get a compact item label and value for debug
            def _fmt_item(idx, col):
                try:
                    lid = df.at[idx, part_col]
                    rid = df.at[idx, "id"] if "id" in df.columns else None
                    val = _coerce_number_like(df.at[idx, col])
                    if val is None:
                        return None
                    tag = str(lid)
                    tag = tag if len(tag) <= 60 else (tag[:57] + "...")
                    if rid:
                        return f"[{rid}] {tag}={val}"
                    return f"{tag}={val}"
                except Exception:
                    return None
            # Coerce numbers for currently selected num_cols (typically c1..cN)
            for c in list(num_cols):
                coerced = df[c].apply(_coerce_number_like)
                if sum(v is not None for v in coerced) == 0:
                    num_cols.remove(c)
                    continue
                df[c] = coerced

            # Fallback: if nothing numeric left (e.g., period columns were renamed and c3/c4 are empty),
            # pick any non-meta columns that contain numbers and coerce them.
            if not num_cols:
                meta_cols = {part_col, "sr_no", "id", "row_type", "parent_id", "components", "calculation_references", "section"}
                for c in [col for col in df.columns if str(col) not in meta_cols]:
                    try:
                        coerced = df[c].apply(_coerce_number_like)
                        if sum(v is not None for v in coerced) > 0:
                            df[c] = coerced
                            num_cols.append(c)
                    except Exception:
                        continue
            try:
                vprint({
                    "phase": "BS.numeric_columns.final",
                    "sheet": sheet_label,
                    "columns": list(map(str, num_cols)),
                })
            except Exception:
                pass
            if debug_enabled():
                try:
                    dprint(f"validation: using numeric columns={num_cols}")
                except Exception:
                    pass
            if not num_cols:
                return

            # Load profile by company type
            ct = _pick_company_type_from_routing(routing_used)
            profile = (VALIDATION_PROFILES.get(ct) or VALIDATION_PROFILES.get("corporate") or {})
            tol_pct = float(profile.get("tolerance_pct", VALIDATION_SUM_TOL_PCT))
            abs_min = float(profile.get("abs_min", VALIDATION_SUM_ABS_MIN))
            excl_pats = _compile(profile.get("exclude_from_sum_patterns", []))
            # Maps present in profile; currently not used here
            eq_checks    = profile.get("equality_checks", []) or []
            try:
                vprint({
                    "phase": "BS.profile",
                    "sheet": sheet_label,
                    "company_type": ct,
                    "tol_pct": tol_pct,
                    "abs_min": abs_min,
                    "exclude_patterns": list(map(str, profile.get("exclude_from_sum_patterns", []) or [])),
                    "eq_checks": eq_checks,
                })
            except Exception:
                pass

            labels = df[part_col].astype(str)
            import re as _re
            total_flag = labels.str.contains(r"^\s*total\b", case=False, regex=True, na=False)
            # Also treat rows explicitly tagged via row_type as totals, even if the text doesn't start with 'Total'
            try:
                type_series_all = df["row_type"].astype(str).str.strip().str.lower()
                rowtype_total_flag = type_series_all.isin(["total","grand_total","subtotal"])  # use as section boundaries
            except Exception:
                type_series_all = None
                rowtype_total_flag = None

            # Helper: get first row index matching a single regex
            def _first_idx(pat_str):
                try:
                    pat = _re.compile(pat_str, _re.I)
                except Exception:
                    return None
                for idx, s in labels.items():
                    if pat.search(str(s)):
                        return idx
                return None

            handled_total_idxs = set()

            # 0) Bank-style sections: validate totals per 'section' when present
            try:
                if VAL_SECTION_CHECKS and ("section" in df.columns):
                    # Normalise row_type for reliable matching
                    try:
                        _rtype = df["row_type"].astype(str).str.strip().str.lower()
                    except Exception:
                        _rtype = df.get("row_type")
                    totals_set = {"total", "grand_total", "subtotal"}
                    line_set   = {"line_item", "item", "line"}
                    for sec in [s for s in df["section"].dropna().unique().tolist() if str(s).strip() != ""]:
                        sec_mask = df["section"].astype(str) == str(sec)
                        sec_df = df.loc[sec_mask]
                        if sec_df.empty:
                            continue
                        # Find the section's declared total row
                        t_idx = None
                        if _rtype is not None:
                            try:
                                t_idx = next((i for i in sec_df.index if str(_rtype.get(i, "")).lower() in totals_set), None)
                            except Exception:
                                t_idx = None
                        if t_idx is None:
                            # fallback: textual 'Total'
                            try:
                                t_idx = next((i for i in sec_df.index if _re.search(r"^\s*total\b", str(sec_df.at[i, part_col]), _re.I)), None)
                            except Exception:
                                t_idx = None
                        if t_idx is None:
                            continue
                        # Item rows in the section
                        if _rtype is not None:
                            item_idxs = [i for i in sec_df.index if str(_rtype.get(i, "")).lower() in line_set]
                        else:
                            # Fallback: anything not 'Total' text
                            item_idxs = [i for i in sec_df.index if not _re.search(r"^\s*total\b", str(sec_df.at[i, part_col] or ""), _re.I)]
                        if not item_idxs:
                            continue
                        for c in num_cols:
                            s = sum((_coerce_number_like(df.at[i, c]) or 0) for i in item_idxs)
                            r = _coerce_number_like(df.at[t_idx, c])
                            if r is None:
                                continue
                            diff = float(r) - float(s)
                            tol = max(abs_min, abs(float(r)) * tol_pct)
                            ok = abs(diff) <= tol
                            details = "; ".join([d for d in (_fmt_item(i, c) for i in item_idxs) if d])
                            part_name = str(df.at[t_idx, part_col])
                            validation_rows.append([sheet_label, part_name, f"Section={sec}", str(c), float(s), float(r), diff, "OK" if ok else "MISMATCH", tol, details])
                            try:
                                vprint({
                                    "phase": "BS.section_check",
                                    "sheet": sheet_label,
                                    "section": str(sec),
                                    "column": str(c),
                                    "sum": float(s),
                                    "reported": float(r),
                                    "diff": diff,
                                    "tol": tol,
                                    "status": "OK" if ok else "MISMATCH",
                                })
                            except Exception:
                                pass
                        handled_total_idxs.add(t_idx)
            except Exception as _e:
                vprint(f"Section-based validation failed: {_e}")

            # 1) Composed totals (e.g., Total equity = Equity attributable + NCI)
            for total_pat, comp_list in (profile.get("composed_totals", {}) or {}).items() if VAL_COMPOSED_AND_GRAND else []:
                t_idx = _first_idx(total_pat)
                if t_idx is None:
                    continue
                comp_idxs = []
                pats = _compile(comp_list)
                comp_idxs = _find_rows_matching(labels, pats)
                if not comp_idxs:
                    continue
                for c in num_cols:
                    s = sum((_coerce_number_like(df.at[i, c]) or 0) for i in comp_idxs)
                    r = _coerce_number_like(df.at[t_idx, c])
                    if r is None:
                        continue
                    diff = float(r) - float(s)
                    tol = max(abs_min, abs(float(r)) * tol_pct)
                    ok = abs(diff) <= tol
                    details = "; ".join([d for d in (_fmt_item(i, c) for i in comp_idxs) if d])
                    part_name = str(df.at[t_idx, part_col])
                    validation_rows.append([sheet_label, part_name, "Composed", str(c), float(s), float(r), diff, "OK" if ok else "MISMATCH", tol, details])
                    try:
                        vprint({
                            "phase": "BS.composed_total",
                            "sheet": sheet_label,
                            "pattern": str(total_pat),
                            "column": str(c),
                            "sum": float(s),
                            "reported": float(r),
                            "diff": diff,
                            "tol": tol,
                            "status": "OK" if ok else "MISMATCH",
                        })
                    except Exception:
                        pass

            # 1b) Declared components on totals/subtotals (explicit calculation_references/components)
            # This is the primary Balance Sheet validation path when configured.
            try:
                if VAL_DECLARED_COMPONENTS and (
                    ("id" in df.columns) and ("row_type" in df.columns) and ("calculation_references" in df.columns or "components" in df.columns)
                ):
                    import json as _json
                    # Build id -> row index map
                    id_map: dict[str, int] = {}
                    for i in df.index:
                        _id = df.at[i, "id"] if "id" in df.columns else None
                        if _id is None or str(_id).strip() == "":
                            continue
                        id_map[str(_id).strip()] = i

                    # Build parent graph: parent_id -> [child_id]
                    children_of: dict[str, list[str]] = {}
                    if "parent_id" in df.columns:
                        for i in df.index:
                            try:
                                pid = df.at[i, "parent_id"]
                                cid = df.at[i, "id"] if "id" in df.columns else None
                                pid_s = str(pid).strip()
                                cid_s = str(cid).strip() if cid is not None else None
                                if pid_s and cid_s:
                                    children_of.setdefault(pid_s, []).append(cid_s)
                            except Exception:
                                continue

                    # Normalise row_type labels
                    try:
                        type_series = df["row_type"].astype(str).str.strip().str.lower()
                        type_series = type_series.replace({
                            "sub_total": "subtotal",
                            "line_item": "line",
                            "lineitem": "line",
                            "grand total": "total",
                            "grandtotal": "total",
                        })
                    except Exception:
                        type_series = df.get("row_type")

                    def _parse_components(val):
                        if val is None or (isinstance(val, float) and _pd.isna(val)):
                            return []
                        if isinstance(val, list):
                            return [str(x).strip() for x in val if str(x).strip()]
                        s = str(val).strip()
                        if s == "":
                            return []
                        # Try JSON array first
                        try:
                            arr = _json.loads(s)
                            if isinstance(arr, list):
                                return [str(x).strip() for x in arr if str(x).strip()]
                        except Exception:
                            pass
                        out = []
                        for tok in _re.split(r"[,|;]", s):
                            tok = tok.strip()
                            tok = _re.sub(r"^[\[\]\s\"\']+|[\[\]\s\"\']+$", "", tok)
                            if tok:
                                out.append(tok)
                        return out

                    # Expand a list of component IDs into leaf row indices (descend into children when needed)
                    def _expand_to_leaf_indices(ids: list[str]) -> list[int]:
                        visited: set[str] = set()
                        out_idx: list[int] = []
                        def _walk(_id: str):
                            if not _id or _id in visited:
                                return
                            visited.add(_id)
                            if _id in id_map:
                                idx = id_map[_id]
                                try:
                                    rtype = str(type_series.get(idx, "") if type_series is not None else df.at[idx, "row_type"]).strip().lower()
                                except Exception:
                                    rtype = ""
                                if rtype in {"item", "line"}:
                                    out_idx.append(idx)
                                    return
                            # Not a direct item → descend into children if any
                            for child in (children_of.get(_id, []) or []):
                                _walk(child)
                            # If no explicit children, include the node itself as a leaf (subtotal value)
                            if _id in id_map and id_map[_id] not in out_idx:
                                out_idx.append(id_map[_id])
                        for _id in (ids or []):
                            _walk(str(_id).strip())
                        # unique preserving order
                        seen = set(); uniq = []
                        for i in out_idx:
                            if i not in seen:
                                uniq.append(i); seen.add(i)
                        return uniq

                    # Prefer more specific section label if available (module-level helper)

                    comp_col = "calculation_references" if "calculation_references" in df.columns else ("components" if "components" in df.columns else None)
                    if comp_col is not None:
                        tot_mask = df.get(comp_col).apply(lambda x: str(x).strip() != "")
                        for t_idx in list(df.index[tot_mask]):
                            comp_ids = _parse_components(df.at[t_idx, comp_col]) if comp_col else []
                            comp_idxs = _expand_to_leaf_indices(comp_ids) if comp_ids else []
                            if not comp_idxs:
                                continue
                            part_name = str(df.at[t_idx, part_col])
                            sec_label = _section_label_from_df(df, t_idx, "Declared")
                            for c in num_cols:
                                s = sum((_coerce_number_like(df.at[i, c]) or 0) for i in comp_idxs)
                                r = _coerce_number_like(df.at[t_idx, c])
                                if r is None:
                                    continue
                                diff = float(r) - float(s)
                                tol = max(abs_min, abs(float(r)) * tol_pct)
                                ok = abs(diff) <= tol
                                details = "; ".join([d for d in (_fmt_item(i, c) for i in comp_idxs) if d])
                                validation_rows.append([sheet_label, part_name, sec_label, str(c), float(s), float(r), diff, "OK" if ok else "MISMATCH", tol, details])
                            handled_total_idxs.add(t_idx)
            except Exception as _e:
                print(f"[Excel] WARN: declared-components validation failed: {_e}", flush=True)

        def _compute_checks_for_cf(sheet_label: str, df_in: "pd.DataFrame") -> None:
            """Cashflow checks: CFO + CFI + CFF ≈ Net Change; Opening + Net Change ≈ Closing."""
            vcfg = (CFG.get("validation", {}) or {}).get("checks", {}) or {}
            cfc = (vcfg.get("cashflow") or {})
            enable = bool(cfc.get("enable", True)) and bool(cfc.get("enforce_cash_tie", True))
            if not enable:
                return
            tol_pct = float(cfc.get("tolerance_pct", (vcfg.get("profit_and_loss") or {}).get("tolerance_pct", 0.005)))
            abs_min = float(cfc.get("abs_min", (vcfg.get("profit_and_loss") or {}).get("abs_min", 1.0)))

            import re as _re
            df = df_in.copy()
            part_col = next((c for c in df.columns if str(c).strip().lower() in {"particulars","particular","description","line item","line_item"}), None)
            if part_col is None:
                return
            # Prefer period-labeled numeric columns over raw cN placeholders.
            meta_cols = {part_col, "sr_no", "id", "row_type", "parent_id", "components", "calculation_references", "section"}
            c_cols   = [c for c in df.columns if _re.fullmatch(r"(?i)c\d+", str(c))]
            other    = [c for c in df.columns if (str(c) not in meta_cols and not _re.fullmatch(r"(?i)c\d+", str(c)))]

            def _numeric_cols(cols):
                out = []
                for c in cols:
                    coerced = df[c].apply(_coerce_number_like)
                    if sum(v is not None for v in coerced) > 0:
                        df[c] = coerced
                        out.append(c)
                return out

            other_num = _numeric_cols(other)
            c_num     = _numeric_cols(c_cols)
            num_cols  = other_num if other_num else c_num
            try:
                print(f"[Validation] CF numeric columns for '{sheet_label}': picked={num_cols} other_num={other_num} c_num={c_num}", flush=True)
            except Exception:
                pass
            if not num_cols:
                return
            try:
                vprint({
                    "phase": "CF.validate.enter",
                    "sheet": sheet_label,
                    "shape": list(getattr(df_in, "shape", [])),
                    "tol_pct": tol_pct,
                    "abs_min": abs_min,
                })
            except Exception:
                pass

            labels = df[part_col].astype(str)
            op_pats = _compile([
                r"^\s*net\s+cash(?:\s+flow)?\b.*operating\s+activities",
                r"^\s*net\s+cash\s+generated.*operating\s+activities",
                r"^\s*net\s+cash\s+from.*operating\s+activities",
                r"^\s*net\s+cash\s+used.*operating\s+activities",
            ])
            inv_pats = _compile([
                r"^\s*net\s+cash(?:\s+flow)?\b.*investing\s+activities",
                r"^\s*net\s+cash\s+used.*investing\s+activities",
                r"^\s*net\s+cash\s+from.*investing\s+activities",
            ])
            fin_pats = _compile([
                r"^\s*net\s+cash(?:\s+flow)?\b.*financing\s+activities",
                r"^\s*net\s+cash\s+used.*financing\s+activities",
                r"^\s*net\s+cash\s+from.*financing\s+activities",
            ])
            net_pats = _compile([
                r"net\s+(?:increase|decrease).*cash\s*(?:and|&)\s*cash\s*equivalents",
                r"net\s+change\s+in\s+cash\s*(?:and|&)\s*cash\s*equivalents",
            ])
            open_pats = _compile([
                r"cash\s*(?:and|&)\s*cash\s*equivalents.*beginning",
                r"opening\s+cash.*equivalents",
            ])
            close_pats = _compile([
                r"cash\s*(?:and|&)\s*cash\s*equivalents.*end",
                r"closing\s+cash.*equivalents",
            ])

            op_rows = _find_rows_matching(labels, op_pats)
            inv_rows = _find_rows_matching(labels, inv_pats)
            fin_rows = _find_rows_matching(labels, fin_pats)
            net_rows = _find_rows_matching(labels, net_pats)
            open_rows = _find_rows_matching(labels, open_pats)
            close_rows = _find_rows_matching(labels, close_pats)
            try:
                print(f"[VALDBG] {{'phase':'CF.patterns.mem','sheet':{sheet_label!r},'op':{op_rows},'inv':{inv_rows},'fin':{fin_rows},'net':{net_rows},'open':{open_rows},'close':{close_rows}}}", flush=True)
            except Exception:
                pass
            try:
                vprint({
                    "phase": "CF.patterns",
                    "sheet": sheet_label,
                    "op_rows": list(map(int, op_rows)),
                    "inv_rows": list(map(int, inv_rows)),
                    "fin_rows": list(map(int, fin_rows)),
                    "net_rows": list(map(int, net_rows)),
                    "open_rows": list(map(int, open_rows)),
                    "close_rows": list(map(int, close_rows)),
                })
            except Exception:
                pass

            # use module-level first-value helper
            def _first_label(rows_idx):
                for i in rows_idx or []:
                    try:
                        s = str(df.at[i, part_col])
                        if s.strip() != "":
                            return s
                    except Exception:
                        continue
                return None

            # Prefer using the user-facing labels for target rows in "Particulars"
            net_label   = _first_label(net_rows) or "Net change in cash and cash equivalents"
            close_label = _first_label(close_rows) or "Cash and cash equivalents at the end of the year"

            for c in num_cols:
                op = _first_val_from_rows(df, op_rows, c)
                inv = _first_val_from_rows(df, inv_rows, c)
                fin = _first_val_from_rows(df, fin_rows, c)
                net = _first_val_from_rows(df, net_rows, c)
                opening = _first_val_from_rows(df, open_rows, c)
                closing = _first_val_from_rows(df, close_rows, c)

                if op is not None and inv is not None and fin is not None and net is not None:
                    total = float(op) + float(inv) + float(fin)
                    diff = total - float(net)
                    tol = max(abs_min, abs(float(net)) * tol_pct)
                    ok = abs(diff) <= tol
                    details = f"CFO={op}; CFI={inv}; CFF={fin}; Net={net}"
                    validation_rows.append([sheet_label, net_label, "CFO+CFI+CFF", str(c), float(total), float(net), diff, "OK" if ok else "MISMATCH", tol, details])

                net2 = net
                if net2 is None and (op is not None and inv is not None and fin is not None):
                    net2 = float(op) + float(inv) + float(fin)
                if opening is not None and closing is not None and net2 is not None:
                    est_close = float(opening) + float(net2)
                    diff = est_close - float(closing)
                    tol = max(abs_min, max(abs(float(est_close)), abs(float(closing))) * tol_pct)
                    ok = abs(diff) <= tol
                    details = f"Opening={opening}; Net={net2}; Closing={closing}"
                    validation_rows.append([sheet_label, close_label, "Opening→Closing", str(c), float(est_close), float(closing), diff, "OK" if ok else "MISMATCH", tol, details])
            # Declared components validation for Cashflow: validate any row with non-empty calculation_references/components
            try:
                vcfg = (CFG.get("validation", {}) or {}).get("checks", {}) or {}
                plc = (vcfg.get("profit_and_loss") or {})
                tol_pct_decl = float(plc.get("tolerance_pct", 0.005))
                abs_min_decl = float(plc.get("abs_min", 1.0))
                max_components = int(plc.get("max_components", 12))
                allow_subset = bool(plc.get("allow_subset", True))

                # Build id maps for expansion
                id_map = {}
                if "id" in df.columns:
                    for i in df.index:
                        _id = df.at[i, "id"]
                        if _id is None or str(_id).strip() == "":
                            continue
                        id_map[str(_id).strip()] = i
                children_of: dict[str, list[str]] = {}
                if "parent_id" in df.columns:
                    for i in df.index:
                        pid = df.at[i, "parent_id"]
                        cid = df.at[i, "id"] if "id" in df.columns else None
                        pid_s = str(pid).strip()
                        cid_s = str(cid).strip() if cid is not None else None
                        if pid_s and cid_s:
                            children_of.setdefault(pid_s, []).append(cid_s)

                try:
                    type_series = df["row_type"].astype(str).str.strip().str.lower()
                except Exception:
                    type_series = None

                # use module-level _parse_components_value

                # Expand a list of IDs into leaf item indices (descend into children if needed)
                def _expand_to_items(ids):
                    visited: set[str] = set()
                    out_idx: list[int] = []
                    def _walk(_id: str):
                        if not _id or _id in visited:
                            return
                        visited.add(_id)
                        if _id in id_map:
                            idx = id_map[_id]
                            try:
                                rt = str(type_series.get(idx, "") if type_series is not None else df.at[idx, "row_type"]).strip().lower()
                            except Exception:
                                rt = ""
                            if rt in {"line_item","line","item"}:
                                out_idx.append(idx)
                                return
                        for child in (children_of.get(_id, []) or []):
                            _walk(child)
                        if _id in id_map and id_map[_id] not in out_idx:
                            out_idx.append(id_map[_id])
                    for _id in ids:
                        _walk(str(_id).strip())
                    # unique preserving order
                    seen = set(); uniq=[]
                    for i in out_idx:
                        if i not in seen:
                            seen.add(i); uniq.append(i)
                    return uniq

                # Simple +/- sign solver reused from P&L
                def _solve_signs(vals, target, tol, allow_subset=True):
                    n = len(vals)
                    if n == 0:
                        return None
                    s = sum(vals)
                    if abs(s - target) <= tol:
                        return ([+1]*n, s)
                    import itertools
                    choices = [-1, 0, +1] if allow_subset else [-1, +1]
                    for signs in itertools.product(choices, repeat=n):
                        if all(x == 0 for x in signs):
                            continue
                        total = sum(signs[i]*vals[i] for i in range(n))
                        if abs(total - target) <= tol:
                            return (list(signs), total)
                    return None

                comp_col = None
                if set(["id","calculation_references"]).issubset(set(map(str, df.columns))):
                    comp_col = "calculation_references"
                elif set(["id","components"]).issubset(set(map(str, df.columns))):
                    comp_col = "components"
                if comp_col is not None:
                    comp_mask = df.get(comp_col).apply(lambda x: str(x).strip() != "")
                    for t_idx in list(df.index[comp_mask]):
                        comp_ids = _parse_components_value(df.at[t_idx, comp_col]) or []
                        comp_idxs = _expand_to_items(comp_ids)
                        if not comp_idxs or len(comp_idxs) > max_components:
                            continue
                        try:
                            part_name = str(df.at[t_idx, part_col])
                        except Exception:
                            part_name = ""
                        sec_label = "Declared Components"
                        for c in num_cols:
                            target = _coerce_number_like(df.at[t_idx, c])
                            if target is None:
                                continue
                            vals = [(_coerce_number_like(df.at[i, c]) or 0.0) for i in comp_idxs]
                            tol = max(abs_min_decl, abs(float(target)) * tol_pct_decl)
                            solved = _solve_signs(vals, float(target), tol, allow_subset)
                            if solved is None:
                                continue
                            signs, total = solved
                            items = []
                            for sign, i in zip(signs, comp_idxs):
                                if sign == 0:
                                    continue
                                sgn = "+" if sign > 0 else "-"
                                try:
                                    _id = df.at[i,'id'] if 'id' in df.columns else ''
                                    _lbl = df.at[i, part_col]
                                    _val = _coerce_number_like(df.at[i, c])
                                except Exception:
                                    _id = ''
                                    _lbl = ''
                                    _val = None
                                items.append(f"{sgn} [{_id}] {_lbl}={_val}")
                            details = " ".join(items)
                            validation_rows.append([sheet_label, part_name, sec_label, str(c), float(total), float(target), float(total) - float(target), "OK", tol, details])
            except Exception as _e:
                print(f"[Excel] WARN: CF declared-components validation failed: {_e}", flush=True)
            # End CF checks
            return

            # 2) Grand totals (e.g., Total assets = Total current + Total non-current)
            for total_pat, comp_list in (profile.get("grand_total_map", {}) or {}).items() if VAL_COMPOSED_AND_GRAND else []:
                t_idx = _first_idx(total_pat)
                if t_idx is None:
                    continue
                comp_idxs = _find_rows_matching(labels, _compile(comp_list))
                if not comp_idxs:
                    continue
                for c in num_cols:
                    s = sum((_coerce_number_like(df.at[i, c]) or 0) for i in comp_idxs)
                    r = _coerce_number_like(df.at[t_idx, c])
                    if r is None:
                        continue
                    diff = float(r) - float(s)
                    tol = max(abs_min, abs(float(r)) * tol_pct)
                    ok = abs(diff) <= tol
                    details = "; ".join([d for d in (_fmt_item(i, c) for i in comp_idxs) if d])
                    part_name = str(df.at[t_idx, part_col])
                    validation_rows.append([sheet_label, part_name, "Grand Total", str(c), float(s), float(r), diff, "OK" if ok else "MISMATCH", tol, details])
                    try:
                        vprint({
                            "phase": "BS.grand_total",
                            "sheet": sheet_label,
                            "pattern": str(total_pat),
                            "column": str(c),
                            "sum": float(s),
                            "reported": float(r),
                            "diff": diff,
                            "tol": tol,
                            "status": "OK" if ok else "MISMATCH",
                        })
                    except Exception:
                        pass
                handled_total_idxs.add(t_idx)

            # 3) Equality checks
            for pair in (eq_checks if VAL_COMPOSED_AND_GRAND else []):
                if not isinstance(pair, (list, tuple)) or len(pair) != 2:
                    continue
                a_idx, b_idx = _first_idx(pair[0]), _first_idx(pair[1])
                if a_idx is None or b_idx is None:
                    continue
                for c in num_cols:
                    a = _coerce_number_like(df.at[a_idx, c])
                    b = _coerce_number_like(df.at[b_idx, c])
                    if a is None or b is None:
                        continue
                    diff = float(a) - float(b)
                    tol = max(abs_min, max(abs(float(a)), abs(float(b))) * tol_pct)
                    ok = abs(diff) <= tol
                    la = str(df.at[a_idx, part_col])
                    lb = str(df.at[b_idx, part_col])
                    details = f"left:{la}={a}; right:{lb}={b}"
                    validation_rows.append([sheet_label, la, "Equality", str(c), float(a), float(b), diff, "OK" if ok else "MISMATCH", tol, details])
                    try:
                        vprint({
                            "phase": "BS.equality",
                            "sheet": sheet_label,
                            "left": la,
                            "right": lb,
                            "column": str(c),
                            "left_val": float(a),
                            "right_val": float(b),
                            "diff": diff,
                            "tol": tol,
                            "status": "OK" if ok else "MISMATCH",
                        })
                    except Exception:
                        pass

            # 3b) Bank-style section equality: if exactly two sections exist, compare their totals
            try:
                if VAL_SECTION_EQUALITY and ("section" in df.columns):
                    secs = [s for s in df["section"].dropna().unique().tolist() if str(s).strip() != ""]
                    if len(secs) == 2:
                        # Determine each section's total row index
                        sec_totals = {}
                        for sec in secs:
                            sec_df = df[df["section"].astype(str) == str(sec)]
                            t_idx = None
                            # Prefer row_type marker
                            try:
                                types = df["row_type"].astype(str).str.strip().str.lower()
                                t_idx = next((i for i in sec_df.index if types.get(i, "") in {"total","grand_total","subtotal"}), None)
                            except Exception:
                                t_idx = None
                            if t_idx is None:
                                # Fallback: textual Total
                                try:
                                    t_idx = next((i for i in sec_df.index if _re.search(r"^\s*total\b", str(sec_df.at[i, part_col] or ""), _re.I)), None)
                                except Exception:
                                    t_idx = None
                            if t_idx is not None:
                                sec_totals[str(sec)] = t_idx
                        if len(sec_totals) == 2:
                            sec_names = list(sec_totals.keys())
                            a_sec, b_sec = sec_names[0], sec_names[1]
                            a_idx, b_idx = sec_totals[a_sec], sec_totals[b_sec]
                            for c in num_cols:
                                a = _coerce_number_like(df.at[a_idx, c])
                                b = _coerce_number_like(df.at[b_idx, c])
                                if a is None or b is None:
                                    continue
                                diff = float(a) - float(b)
                                tol = max(abs_min, max(abs(float(a)), abs(float(b))) * tol_pct)
                                ok = abs(diff) <= tol
                                details = f"left:{a_sec}={a}; right:{b_sec}={b}"
                                part_name = str(df.at[a_idx, part_col])
                                validation_rows.append([sheet_label, part_name, "Section Equality", str(c), float(a), float(b), diff, "OK" if ok else "MISMATCH", tol, details])
                                try:
                                    vprint({
                                        "phase": "BS.section_equality",
                                        "sheet": sheet_label,
                                        "left_section": a_sec,
                                        "right_section": b_sec,
                                        "column": str(c),
                                        "left_val": float(a),
                                        "right_val": float(b),
                                        "diff": diff,
                                        "tol": tol,
                                        "status": "OK" if ok else "MISMATCH",
                                    })
                                except Exception:
                                    pass
            except Exception as _e:
                vprint(f"Section-equality validation failed: {_e}")

            # 4) Declared components in DataFrame (row_type=subtotal/grand_total with calculation_references/components)
            try:
                if VAL_DECLARED_COMPONENTS and (
                    set(["id","row_type","calculation_references"]).issubset(set(map(str, df.columns)))
                    or set(["id","row_type","components"]).issubset(set(map(str, df.columns)))
                ):
                    import json as _json
                    # Build id -> row index map (case-insensitive ids)
                    id_map = {}
                    for i in df.index:
                        _id = df.at[i, "id"] if "id" in df.columns else None
                        if _id is None or str(_id).strip() == "":
                            continue
                        id_map[str(_id).strip()] = i

                    # Build parent_id -> [child_ids] map to support tree expansion
                    children_of: dict[str, list[str]] = {}
                    if "parent_id" in df.columns:
                        for i in df.index:
                            try:
                                pid = df.at[i, "parent_id"]
                                cid = df.at[i, "id"] if "id" in df.columns else None
                                pid_s = str(pid).strip()
                                cid_s = str(cid).strip() if cid is not None else None
                                if pid_s and cid_s:
                                    children_of.setdefault(pid_s, []).append(cid_s)
                            except Exception:
                                continue

                    # use module-level _parse_components_value

                    # Normalise row_type synonyms
                    try:
                        type_series = df["row_type"].astype(str).str.strip().str.lower()
                        type_series = type_series.replace({
                            "sub_total": "subtotal",
                            "line_item": "line",
                            "lineitem": "line",
                            "grand total": "total",
                            "grandtotal": "total",
                        })
                    except Exception:
                        type_series = df["row_type"].astype(str).str.lower()
                    comp_col = "calculation_references" if "calculation_references" in df.columns else "components"
                    # Validate ALL rows that declare components, irrespective of row_type
                    tot_mask = df.get(comp_col).apply(lambda x: str(x).strip() != "")

                    # Helper: recursively expand a list of component IDs into leaf row indices.
                    # If an ID points to a heading or subtotal, include its descendants by parent_id.
                    def _expand_to_leaf_indices(ids: list[str]) -> list[int]:
                        visited: set[str] = set()
                        out_idx: list[int] = []

                        def _walk(_id: str):
                            if not _id or _id in visited:
                                return
                            visited.add(_id)
                            if _id in id_map:
                                idx = id_map[_id]
                                try:
                                    rtype = str(type_series.get(idx, "")).strip().lower()
                                except Exception:
                                    rtype = ""
                                if rtype == "item":
                                    out_idx.append(idx)
                                    return
                            # Not a direct item → descend to children if any
                            children = (children_of.get(_id, []) or [])
                            if children:
                                for child in children:
                                    _walk(child)
                            else:
                                # No explicit children in parent_id graph → include the node itself as a leaf (subtotal value)
                                if _id in id_map:
                                    out_idx.append(id_map[_id])

                        for _id in ids:
                            _walk(str(_id).strip())
                        # Unique while preserving order
                        seen = set()
                        uniq = []
                        for i in out_idx:
                            if i not in seen:
                                uniq.append(i)
                                seen.add(i)
                        return uniq
                    # Helper to fetch label for 'Section' column: use module-level helper

                    comp_col = "calculation_references" if "calculation_references" in df.columns else ("components" if "components" in df.columns else None)
                    for t_idx in list(df.index[tot_mask]):
                        comp_ids = _parse_components_value(df.at[t_idx, comp_col]) if comp_col else []
                        # Expand components via parent-child tree when needed (headings/subtotals)
                        comp_idxs = _expand_to_leaf_indices(comp_ids) if comp_ids else []
                        if valdbg_enabled():
                            missing_ids = [cid for cid in comp_ids if cid not in id_map]
                            if missing_ids:
                                vprint(f"Declared components not found in id_map (sheet={sheet_label}, row={t_idx}):", missing_ids)
                        if not comp_idxs:
                            continue
                        sec_label = _section_label_from_df(df, t_idx, "Declared")
                        part_name = str(df.at[t_idx, part_col])
                        for c in num_cols:
                            s = sum((_coerce_number_like(df.at[i, c]) or 0) for i in comp_idxs)
                            r = _coerce_number_like(df.at[t_idx, c])
                            if r is None:
                                continue
                            diff = float(r) - float(s)
                            tol = max(abs_min, abs(float(r)) * tol_pct)
                            ok = abs(diff) <= tol
                            details = "; ".join([d for d in (_fmt_item(i, c) for i in comp_idxs) if d])
                            part_name = str(df.at[t_idx, part_col])
                            validation_rows.append([sheet_label, part_name, sec_label, str(c), float(s), float(r), diff, "OK" if ok else "MISMATCH", tol, details])
                        handled_total_idxs.add(t_idx)

                    # Also validate totals that have no explicit components but have children rows via parent_id
                    try:
                        if VAL_CHILDREN_WITHOUT_COMPONENTS and "id" in df.columns and "parent_id" in df.columns:
                            comp_col = "calculation_references" if "calculation_references" in df.columns else ("components" if "components" in df.columns else None)
                            for t_idx in list(df.index[type_series.isin(["subtotal","grand_total","total"]) & ((df.get(comp_col).astype(str).str.strip() == "") if comp_col else False) ]):
                                try:
                                    tid = str(df.at[t_idx, "id"]).strip()
                                except Exception:
                                    tid = ""
                                if not tid or tid not in children_of:
                                    continue
                                comp_idxs = _expand_to_leaf_indices(children_of.get(tid, []) or [])
                                if not comp_idxs:
                                    continue
                                for c in num_cols:
                                    s = sum((_coerce_number_like(df.at[i, c]) or 0) for i in comp_idxs)
                                    r = _coerce_number_like(df.at[t_idx, c])
                                    if r is None:
                                        continue
                                    diff = float(r) - float(s)
                                    tol = max(abs_min, abs(float(r)) * tol_pct)
                                    ok = abs(diff) <= tol
                                    details = "; ".join([d for d in (_fmt_item(i, c) for i in comp_idxs) if d])
                                    part_name = str(df.at[t_idx, part_col])
                                    validation_rows.append([sheet_label, part_name, "Children", str(c), float(s), float(r), diff, "OK" if ok else "MISMATCH", tol, details])
                                    try:
                                        vprint({
                                            "phase": "BS.children_without_components",
                                            "sheet": sheet_label,
                                            "total_row": part_name,
                                            "column": str(c),
                                            "sum": float(s),
                                            "reported": float(r),
                                            "diff": diff,
                                            "tol": tol,
                                            "status": "OK" if ok else "MISMATCH",
                                            "n_items": len(comp_idxs),
                                        })
                                    except Exception:
                                        pass
                                handled_total_idxs.add(t_idx)
                    except Exception:
                        pass
            except Exception as _e:
                print(f"[Excel] WARN: declared-components validation failed: {_e}", flush=True)

            # 5) Fallback contiguous blocks for remaining 'Total ...' rows
            start = 0
            section_no = 0
            total_idxs = list(df.index[total_flag])
            if rowtype_total_flag is not None:
                try:
                    total_idxs = sorted(set(total_idxs) | set(list(df.index[rowtype_total_flag])))
                except Exception:
                    pass
            for t_idx in (total_idxs if VAL_BLOCK_FALLBACK else []):
                if t_idx in handled_total_idxs:
                    start = t_idx + 1
                    continue
                end = t_idx
                section = df.iloc[start:end]
                # Exclude any rows matching exclude patterns (subtotals etc.)
                if excl_pats:
                    mask_excl = section[part_col].astype(str).apply(lambda s: any(p.search(s) for p in excl_pats))
                    section = section.loc[~mask_excl]
                # Detect nested subtotals inside the section to avoid double-counting
                try:
                    tol_for = lambda r: max(abs_min, abs(float(r)) * tol_pct) if r is not None else abs_min
                    # Numeric row indices inside section (has at least one numeric value)
                    numeric_rows = [i for i in section.index if any(_coerce_number_like(section.at[i, c]) is not None for c in num_cols)]
                    internal_subtotals = set()
                    g_start = 0
                    for pos in range(len(numeric_rows)):
                        idx = numeric_rows[pos]
                        if pos == g_start:
                            continue
                        prev_idxs = numeric_rows[g_start:pos]
                        if not prev_idxs:
                            continue
                        # Check if current row equals sum of previous contiguous block (all available columns)
                        is_sub = True
                        for c in num_cols:
                            r = _coerce_number_like(section.at[idx, c])
                            if r is None:
                                continue
                            s_val = sum((_coerce_number_like(section.at[j, c]) or 0) for j in prev_idxs)
                            if abs(float(r) - float(s_val)) > tol_for(r):
                                is_sub = False
                                break
                        if is_sub:
                            internal_subtotals.add(idx)
                            g_start = pos + 1  # reset after subtotal
                    if internal_subtotals:
                        section = section.loc[~section.index.isin(internal_subtotals)]
                except Exception as _e:
                    print(f"[Excel] WARN: nested subtotal detection failed: {_e}", flush=True)
                if section.shape[0] == 0:
                    start = t_idx + 1
                    continue
                section_no += 1
                label = f"Section {section_no}"
                for c in num_cols:
                    s = _pd.to_numeric(section[c], errors="coerce").fillna(0).sum()
                    r = _coerce_number_like(df.at[t_idx, c])
                    if r is None:
                        continue
                    diff = float(r) - float(s)
                    tol = max(abs_min, abs(float(r)) * tol_pct)
                    ok = abs(diff) <= tol
                    # details: each included line with numeric value for this column
                    comp_idxs = [i for i in section.index if _coerce_number_like(section.at[i, c]) is not None]
                    details = "; ".join([d for d in (_fmt_item(i, c) for i in comp_idxs) if d])
                    part_name = str(df.at[t_idx, part_col])
                    validation_rows.append([sheet_label, part_name, label, str(c), float(s), float(r), diff, "OK" if ok else "MISMATCH", tol, details])
                    try:
                        vprint({
                            "phase": "BS.block_fallback",
                            "sheet": sheet_label,
                            "block_label": label,
                            "column": str(c),
                            "sum": float(s),
                            "reported": float(r),
                            "diff": diff,
                            "tol": tol,
                            "status": "OK" if ok else "MISMATCH",
                            "n_items": len(comp_idxs),
                        })
                    except Exception:
                        pass
                start = t_idx + 1
        
        def _compute_checks_for_pl(sheet_label: str, df_in: "pd.DataFrame") -> None:
            cfg = (CFG.get("validation", {}) or {}).get("checks", {}) or {}
            plc = (cfg.get("profit_and_loss") or {})
            if not bool(plc.get("enable", True)):
                return
            tol_pct = float(plc.get("tolerance_pct", 0.005))
            abs_min = float(plc.get("abs_min", 1.0))
            use_declared_first = bool(plc.get("use_declared_components_first", True))
            declared_only = bool(plc.get("declared_only", True))
            max_components = int(plc.get("max_components", 12))
            allow_subset = bool(plc.get("allow_subset", False))
            try:
                vprint({
                    "phase": "PL.validate.enter",
                    "sheet": sheet_label,
                    "shape": list(getattr(df_in, "shape", [])),
                    "tol_pct": tol_pct,
                    "abs_min": abs_min,
                    "use_declared_first": use_declared_first,
                    "declared_only": declared_only,
                    "max_components": max_components,
                    "allow_subset": allow_subset,
                })
            except Exception:
                pass

            import re as _re
            df = df_in.copy()
            part_col = next((c for c in df.columns if str(c).strip().lower() in {"particulars","particular","description","line item","line_item"}), None)
            if part_col is None:
                return
            # numeric columns
            num_cols = [c for c in df.columns if _re.fullmatch(r"(?i)c\d+", str(c))]
            if not num_cols:
                meta_cols = {part_col, "sr_no", "id", "row_type", "parent_id", "components", "calculation_references", "section"}
                num_cols = [c for c in df.columns if str(c) not in meta_cols]
            # coerce
            for c in list(num_cols):
                coerced = df[c].apply(_coerce_number_like)
                if sum(v is not None for v in coerced) == 0:
                    num_cols.remove(c)
                    continue
                df[c] = coerced
            if not num_cols:
                return
            try:
                vprint({
                    "phase": "PL.numeric_columns.final",
                    "sheet": sheet_label,
                    "columns": list(map(str, num_cols)),
                })
            except Exception:
                pass
            # Quick debug: show numeric columns picked
            try:
                print(f"[Validation] PL numeric columns for '{sheet_label}': {num_cols}", flush=True)
            except Exception:
                pass

            # id/parent mapping for declared
            id_map = {}
            if "id" in df.columns:
                for i in df.index:
                    _id = df.at[i, "id"]
                    if _id is None or str(_id).strip() == "":
                        continue
                    id_map[str(_id).strip()] = i
            children_of: dict[str, list[str]] = {}
            if "parent_id" in df.columns:
                for i in df.index:
                    pid = df.at[i, "parent_id"]
                    cid = df.at[i, "id"] if "id" in df.columns else None
                    pid_s = str(pid).strip()
                    cid_s = str(cid).strip() if cid is not None else None
                    if pid_s and cid_s:
                        children_of.setdefault(pid_s, []).append(cid_s)

            try:
                type_series = df["row_type"].astype(str).str.strip().str.lower()
                type_series = type_series.replace({
                    "sub_total": "subtotal",
                    "lineitem": "line",
                    "line_item": "line",
                    "grand total": "total",
                    "grandtotal": "total",
                })
            except Exception:
                type_series = df.get("row_type")

            def _expand_to_items(ids: list[str]) -> list[int]:
                visited: set[str] = set()
                out_idx: list[int] = []
                def _walk(_id: str):
                    if not _id or _id in visited:
                        return
                    visited.add(_id)
                    if _id in id_map:
                        idx = id_map[_id]
                        try:
                            rtype = str(type_series.get(idx, "")).strip().lower()
                        except Exception:
                            rtype = ""
                        if rtype == "item" or rtype == "line":
                            out_idx.append(idx)
                            return
                    kids = (children_of.get(_id, []) or [])
                    if kids:
                        for ch in kids:
                            _walk(ch)
                    else:
                        # leaf subtotal without explicit children → include its own value
                        if _id in id_map:
                            out_idx.append(id_map[_id])
                for _id in ids:
                    _walk(str(_id).strip())
                # unique preserve order
                seen = set(); uniq = []
                for i in out_idx:
                    if i not in seen:
                        uniq.append(i); seen.add(i)
                return uniq

            # use module-level helper for section label with P&L default

            def _solve_signs(vals: list[float], target: float, tol: float, allow_subset: bool) -> tuple[list[int], float] | None:
                n = len(vals)
                if n == 0:
                    return None
                if n > max_components:
                    return None
                # fast path: all-positive
                s = sum(vals)
                if abs(s - target) <= tol:
                    return ([+1]*n, s)
                # brute-force signs (+/-); optionally allow zero (ignore) when allow_subset
                import itertools
                if allow_subset:
                    choices = [-1, 0, +1]
                else:
                    choices = [-1, +1]
                for signs in itertools.product(choices, repeat=n):
                    if all(x == 0 for x in signs):
                        continue
                    total = sum(signs[i]*vals[i] for i in range(n))
                    if abs(total - target) <= tol:
                        return (list(signs), total)
                return None

            # Declared totals first (recommended)
            # use module-level _parse_components_value

            # Validate ONLY rows that have non-empty calculation_references
            comp_col = None
            if use_declared_first and set(["id","row_type","calculation_references"]).issubset(set(map(str, df.columns))):
                comp_col = "calculation_references"
            if comp_col is not None:
                # Strict: only rows with actual calculation_references content
                def _has_calc_refs(x) -> bool:
                    try:
                        arr = _parse_components_value(x)
                        return bool(arr)
                    except Exception:
                        return False
                tot_mask = df.get(comp_col).apply(_has_calc_refs)
                for t_idx in list(df.index[tot_mask]):
                    comp_ids = _parse_components_value(df.at[t_idx, comp_col]) or []
                    comp_idxs = _expand_to_items(comp_ids)
                    if not comp_idxs:
                        continue
                    sec_label = _section_label_from_df(df, t_idx, "P&L")
                    try:
                        part_name = str(df.at[t_idx, part_col])
                    except Exception:
                        part_name = ""
                    for c in num_cols:
                        target = _coerce_number_like(df.at[t_idx, c])
                        if target is None:
                            continue
                        vals = [(_coerce_number_like(df.at[i, c]) or 0.0) for i in comp_idxs]
                        tol = max(abs_min, abs(float(target)) * tol_pct)
                        solved = _solve_signs(vals, float(target), tol, allow_subset)
                        if solved is None:
                            continue
                        signs, total = solved
                        # Build details/formula
                        items = []
                        for sign, i in zip(signs, comp_idxs):
                            if sign == 0:
                                continue
                            sgn = "+" if sign > 0 else "-"
                            items.append(f"{sgn} [{df.at[i,'id']}] {df.at[i, part_col]}={_coerce_number_like(df.at[i, c])}")
                        details = " ".join(items)
                        validation_rows.append([sheet_label, part_name, sec_label, str(c), float(total), float(target), float(total) - float(target), "OK", tol, details])
                        try:
                            vprint({
                                "phase": "PL.declared_components",
                                "sheet": sheet_label,
                                "total_row": part_name,
                                "section": sec_label,
                                "column": str(c),
                                "target": float(target),
                                "computed": float(total),
                                "diff": float(total) - float(target),
                                "tol": tol,
                                "n_components": len(comp_idxs),
                                "signs": signs,
                            })
                        except Exception:
                            pass

            # If only declared references are allowed (default), stop here
            if declared_only:
                return

            # Optional: contiguous block fallback near 'Total ...' lines (same as BS but with sign search)
            labels = df[part_col].astype(str)
            total_flag = labels.str.contains(r"^\s*total\b", case=False, regex=True, na=False)
            start = 0
            total_idxs = list(df.index[total_flag])
            for t_idx in total_idxs:
                end = t_idx
                block = df.iloc[start:end]
                start = t_idx + 1
                if block.empty:
                    continue
                candidates = [i for i in block.index if any(_coerce_number_like(block.at[i, c]) is not None for c in num_cols)]
                if len(candidates) == 0 or len(candidates) > max_components:
                    continue
                sec_label = _section_label_from_df(df, t_idx, "P&L")
                for c in num_cols:
                    target = _coerce_number_like(df.at[t_idx, c])
                    if target is None:
                        continue
                    vals = [(_coerce_number_like(block.at[i, c]) or 0.0) for i in candidates]
                    tol = max(abs_min, abs(float(target)) * tol_pct)
                    solved = _solve_signs(vals, float(target), tol, allow_subset)
                    if solved is None:
                        continue
                    signs, total = solved
                    items = []
                    for sign, i in zip(signs, candidates):
                        if sign == 0:
                            continue
                        sgn = "+" if sign > 0 else "-"
                        items.append(f"{sgn} [{df.at[i,'id'] if 'id' in df.columns else ''}] {block.at[i, part_col]}={_coerce_number_like(block.at[i, c])}")
                    details = " ".join(items)
                    part_name = str(df.at[t_idx, part_col])
                    validation_rows.append([sheet_label, part_name, sec_label, str(c), float(total), float(target), float(total) - float(target), "OK", tol, details])
                    try:
                        vprint({
                            "phase": "PL.block_fallback",
                            "sheet": sheet_label,
                            "total_row": part_name,
                            "section": sec_label,
                            "column": str(c),
                            "target": float(target),
                            "computed": float(total),
                            "diff": float(total) - float(target),
                            "tol": tol,
                            "n_candidates": len(candidates),
                            "signs": signs,
                        })
                    except Exception:
                        pass
        def _use_labels_for_sheet(name: str) -> bool:
            # Accept: bool | {sheet_name: bool, ...}
            val = use_period_labels_cfg
            if isinstance(val, dict):
                # Exact match first
                if name in val:
                    return bool(val[name])
                # Try case-insensitive match
                for k, v in val.items():
                    try:
                        if str(k).strip().lower() == str(name).strip().lower():
                            return bool(v)
                    except Exception:
                        continue
                # Support optional 'default' key
                if "default" in val:
                    try:
                        return bool(val.get("default"))
                    except Exception:
                        return True
                # Fallback default when dict provided but no key: True
                return True
            return bool(val)

    for sheet_name in sheet_order:
        if first_written and first_name == sheet_name:
            # Already wrote this one in the pre-write block
            continue
        _sheet_t0 = time.perf_counter()
        df = combined_sheets.get(sheet_name)
        if df is None or getattr(df, "empty", True):
            df = pd.DataFrame(columns=["Particulars"])
            pre_cols = list(df.columns)
            print(f"[Excel] [{sheet_name}] columns_before={pre_cols}", flush=True)
            print(f"[Excel] [{sheet_name}] is_empty=True rows=0 cols={len(pre_cols)}", flush=True)
        else:
            df = _normalize_df_for_excel(sheet_name, df)
            pre_cols = list(df.columns)
            try:
                _mem = int(df.memory_usage(index=True, deep=True).sum()) if hasattr(df, "memory_usage") else -1
            except Exception:
                _mem = -1
            print(f"[Excel] [{sheet_name}] columns_before={pre_cols}", flush=True)
            try:
                print(f"[Excel] [{sheet_name}] shape(before)={getattr(df, 'shape', (0,0))} approx_mem={_mem}B", flush=True)
            except Exception:
                pass
        try:
            logger.info("Excel pre-rename [%s] cols=%s", sheet_name, list(df.columns))
        except Exception:
            pass
        finally:
            # Optionally rename c1..cN headers to actual period labels (per-sheet toggle)
            if _use_labels_for_sheet(sheet_name):
                try:
                    # 0) Resolve labels for this sheet from (local → global) caches
                    period_labels = _pick_period_labels_for_sheet(
                        sheet_name,
                        period_labels_by_doc,     # local: from *_ocr.json scan or arg
                        PERIOD_LABELS_BY_DOC      # global cache from earlier runs
                    )
                    print(
                        f"[Excel] resolve-period-labels: sheet={sheet_name!r} "
                        f"local_keys={list((period_labels_by_doc.get(sheet_name, {}) or {}).keys()) if isinstance(period_labels_by_doc, dict) else []} "
                        f"global_keys={list((PERIOD_LABELS_BY_DOC.get(sheet_name, {}) or {}).keys()) if isinstance(PERIOD_LABELS_BY_DOC, dict) else []}",
                        flush=True
                    )

                    # Flatten dict-of-dicts ({'c1': {'label': '...'}}) → {'c1': '...'} while preserving order
                    if period_labels and any(isinstance(v, dict) for v in period_labels.values()):
                        period_labels = {str(k).lower(): (v.get("label") or "") if isinstance(v, dict) else str(v)
                                         for k, v in period_labels.items()}
                    else:
                        period_labels = {str(k).lower(): ("" if v is None else str(v))
                                         for k, v in (period_labels or {}).items()}

                    print(f"[Excel] sheet={sheet_name!r} period_label_keys={list(period_labels.keys()) if period_labels else []}")
                    if not period_labels:
                        logger.warning("No period labels found for sheet: %s", sheet_name)
                        print(f"[Excel] WARN: no period labels found for {sheet_name!r} — columns will remain as c1,c2,...", flush=True)
                        # still record sheet in debug dump
                        debug_dump["sheets"][sheet_name] = {
                            "columns_before": pre_cols,
                            "period_label_keys": [],
                            "rename_map": {},
                            "columns_after": list(df.columns),
                        }
                    else:
                        rename_map: dict[str, str] = {}
                        has_indexed_keys = any(re.fullmatch(r"(?i)[cp]\d+", str(k)) for k in period_labels.keys())
                        if has_indexed_keys:
                            # Strategy A: direct key match on cN/pN
                            for col in list(df.columns):
                                low = str(col).strip().lower()
                                m = re.fullmatch(r"([cp])(\d+)", low)
                                if not m:
                                    continue
                                prefix, num = m.group(1), m.group(2)
                                key = f"{prefix}{num}"
                                alt = f"{'p' if prefix == 'c' else 'c'}{num}"
                                label = period_labels.get(key) or period_labels.get(alt)
                                if label:
                                    rename_map[col] = str(label)
                        else:
                            # Strategy B: order-based fallback for non-indexed ids (e.g., 'quarter_ended_...')
                            try:
                                c_cols = sorted(
                                    [c for c in df.columns if re.fullmatch(r"(?i)c\d+", str(c))],
                                    key=lambda x: int(re.findall(r"\d+", str(x))[0])
                                )
                            except Exception:
                                c_cols = [c for c in df.columns if str(c).lower().startswith('c')]
                            labels_in_order = [str(v) for v in period_labels.values() if str(v).strip() != ""]
                            n = min(len(c_cols), len(labels_in_order))
                            for i in range(n):
                                rename_map[c_cols[i]] = labels_in_order[i]

                        logger.info("Excel rename [%s] using labels=%s → map=%s", sheet_name, sorted(period_labels.keys()), rename_map)
                        print(f"[Excel] rename-map for {sheet_name!r}: {rename_map}", flush=True)

                        if rename_map:
                            df = df.rename(columns=rename_map)
                            print(f"[Excel] [{sheet_name}] columns_after={list(df.columns)}", flush=True)

                        # record in debug dump (whether or not a rename happened)
                        debug_dump["sheets"][sheet_name] = {
                            "columns_before": pre_cols,
                            "period_label_keys": sorted(list(period_labels.keys())),
                            "rename_map": rename_map.copy(),
                            "columns_after": list(df.columns),
                        }
                except Exception as e:
                    logger.error("Header rename failed for %s: %s", sheet_name, e)
                    print(f"[Excel] ERROR while renaming headers for {sheet_name!r}: {e}", flush=True)

        # Drop empty period/numeric columns (e.g., blank c3/c4/c5 or empty labeled periods)
        try:
                import re as _re
                meta_cols = {"particulars","sr_no","id","row_type","parent_id","components","calculation_references","section","section_id","sectionid","section_id"}
                drop_cols = []
                for col in list(df.columns):
                    low = str(col).strip().lower()
                    if low in meta_cols:
                        continue
                    # Only consider period-ish columns: cN/pN or non-meta columns likely holding numbers
                    if not (_re.fullmatch(r"(?i)[cp]\\d+", str(col)) or True):
                        # 'or True' keeps behavior simple: evaluate all non-meta columns for numeric emptiness
                        pass
                    coerced = df[col].apply(_coerce_number_like)
                    if sum(v is not None for v in coerced) == 0:
                        drop_cols.append(col)
                if drop_cols:
                    print(f"[Excel] [{sheet_name}] dropping empty columns: {drop_cols}", flush=True)
                    df = df.drop(columns=drop_cols)
        except Exception as _de:
            print(f"[Excel] WARN: drop-empty-columns failed for {sheet_name!r}: {_de}", flush=True)

        safe_name = sheet_name[:31] or "Sheet"
        # Per-sheet dtype summary for debugging
        try:
            _dtypes = [str(df.dtypes.__getitem__(c)) for c in df.columns]
            print(f"[Excel] [{sheet_name}] dtypes(before_write)={dict(zip(df.columns, _dtypes))}", flush=True)
        except Exception:
            pass
        try:
            print(f"[Excel] [{sheet_name}] writing → sheet={safe_name} rows={getattr(df,'shape',(0,0))[0]} cols={getattr(df,'shape',(0,0))[1]}", flush=True)
        except Exception:
            pass
        try:
            df = _sanitize_df_for_excel(df)
            safe_name = _unique_sheet_name(sheet_name)
            df.to_excel(writer, sheet_name=safe_name, index=False)
            wrote_any_real_sheet = True
        except Exception as _we:
            print(f"[Excel] ERROR writing sheet '{sheet_name}' → {safe_name}: {_we}", flush=True)
            # Continue to next sheet; placeholder ensures workbook remains valid
            continue

        ws = writer.book[safe_name]
        try:
            ws.sheet_state = "visible"
        except Exception:
            pass
        header_font  = Font(bold=True, color=header_font_color)
        header_fill  = PatternFill("solid", fgColor=header_fill_hex)
        header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell_align   = Alignment(vertical="top", wrap_text=True)

        max_width = 60
        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
            # autosize
            longest = 0
            for c in col_cells:
                val = "" if c.value is None else str(c.value)
                longest = max(longest, len(val))
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(longest + 2, 10), max_width)

            # header style
            h = col_cells[0]
            h.font = header_font
            h.fill = header_fill
            h.alignment = header_align

            # data cells style + number format if numeric present
            any_num = any(isinstance(c.value, (int, float)) for c in col_cells[1:])
            for c in col_cells[1:]:
                c.alignment = cell_align
                if any_num and (str(h.value).strip().lower() not in {"particulars","description","particular"}):
                    c.number_format = "#,##0.00_);(#,##0.00)"

            try:
                ws_debug_headers = [cell.value for cell in ws[1]]
                logger.info("Excel post-rename [%s] header row=%s", safe_name, ws_debug_headers)
                print(f"[Excel] [{sheet_name}] worksheet dims rows={ws.max_row} cols={ws.max_column}", flush=True)
            except Exception:
                pass

            try:
                ws.freeze_panes = freeze_panes
            except Exception:
                ws.freeze_panes = "A2"

            # Optional Common-Size sheet right after the base sheet
            try:
                if bool((CFG.get("export", {}) or {}).get("statements_workbook", {}).get("include_common_size_sheets", True)):
                    # Rebuild rows from df to feed analytics common-size
                    try:
                        rows_for_cs = []
                        for _, row in df.iterrows():
                            r = {k: row[k] for k in df.columns if k in (list(df.columns))}
                            rows_for_cs.append(r)
                    except Exception:
                        rows_for_cs = []
                    # Build labels map from period_labels_by_doc or global cache
                    labels_map = _pick_period_labels_for_sheet(
                        sheet_name,
                        period_labels_by_doc,
                        PERIOD_LABELS_BY_DOC,
                    ) or {}
                    # Normalize to key->label strings
                    if labels_map and any(isinstance(v, dict) for v in labels_map.values()):
                        labels_map = {str(k).lower(): str((v or {}).get("label", "")) for k, v in (labels_map or {}).items()}
                    else:
                        labels_map = {str(k).lower(): str(v) for k, v in (labels_map or {}).items()}
                    cs = _analytics.compute_common_size(sheet_name, rows_for_cs, labels_map)
                    if cs and (cs.get("rows") or []):
                        import pandas as _pd
                        df_cs = _pd.DataFrame(cs.get("rows") or [])
                        # keep columns order: Particulars, then period columns in detected order
                        cols = [c for c in df_cs.columns if str(c).lower() == "particulars"]
                        try:
                            _pcols = [c for c in df_cs.columns if re.fullmatch(r"(?i)[cp]\\d+", str(c))]
                            # sort c1..cN by index
                            _pcols = sorted(_pcols, key=lambda x: int(re.findall(r"\d+", str(x))[0]))
                        except Exception:
                            _pcols = [c for c in df_cs.columns if c not in cols]
                        ordered = cols + _pcols + [c for c in df_cs.columns if c not in cols + _pcols]
                        df_cs = df_cs.loc[:, ordered]
                        df_cs = _sanitize_df_for_excel(df_cs)
                        safe_cs = _unique_sheet_name(sheet_name + " (Common Size)")
                        df_cs.to_excel(writer, sheet_name=safe_cs, index=False)
                        ws_cs = writer.book[safe_cs]
                        # Style similar to base, with percentage format for numeric columns
                        header_font  = Font(bold=True, color=header_font_color)
                        header_fill  = PatternFill("solid", fgColor=header_fill_hex)
                        header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
                        cell_align   = Alignment(vertical="top", wrap_text=True)
                        for col_cells in ws_cs.iter_cols(min_row=1, max_row=ws_cs.max_row):
                            longest = max(len("" if c.value is None else str(c.value)) for c in col_cells)
                            ws_cs.column_dimensions[col_cells[0].column_letter].width = min(max(longest + 2, 10), 60)
                            h = col_cells[0]
                            h.font = header_font
                            h.fill = header_fill
                            h.alignment = header_align
                            any_num = any(isinstance(c.value, (int, float)) for c in col_cells[1:])
                            for c in col_cells[1:]:
                                c.alignment = cell_align
                                if any_num and (str(h.value).strip().lower() not in {"particulars","description","particular"}):
                                    c.number_format = "0.00%"
                        try:
                            ws_cs.freeze_panes = freeze_panes
                        except Exception:
                            ws_cs.freeze_panes = "A2"
            except Exception as _ec:
                print(f"[Excel] WARN: common-size sheet for {sheet_name!r} failed: {_ec}", flush=True)

            # Collect validation rows for Balance Sheet / P&L / Cashflow sheets
            try:
                name_l = str(sheet_name).strip().lower()
                vcfg = (CFG.get("validation", {}) or {}).get("checks", {}) or {}
                bs_enabled  = bool(((vcfg.get("balance_sheet") or {}).get("sum_subitems") or {}).get("enable", True))
                pl_enabled  = bool((vcfg.get("profit_and_loss") or {}).get("enable", True))
                cf_enabled  = bool((vcfg.get("cashflow") or {}).get("enable", True)) and bool((vcfg.get("cashflow") or {}).get("enforce_cash_tie", True))
                is_bs = ("balance" in name_l and "sheet" in name_l) or ("statement of assets and liabilities" in name_l)
                is_pl = ("profit" in name_l and "loss" in name_l)
                is_cf = ("cash" in name_l and "flow" in name_l)
                will_validate = INCLUDE_VALIDATION_SHEET and ((is_bs and bs_enabled) or (is_pl and pl_enabled) or (is_cf and cf_enabled))
                print(f"[Excel] Validation check: sheet='{sheet_name}', rows={getattr(df, 'shape', [0,0])[0]}, will_validate={will_validate}", flush=True)
                if will_validate:
                    _before = len(validation_rows)
                    if is_bs and bs_enabled:
                        _compute_sum_checks_for_bs(sheet_name, df)
                    elif is_pl and pl_enabled:
                        _compute_checks_for_pl(sheet_name, df)
                    elif is_cf and cf_enabled:
                        _compute_checks_for_cf(sheet_name, df)
                    _added = len(validation_rows) - _before
                    print(f"[Excel] Validation added {_added} row(s) for sheet '{sheet_name}'", flush=True)
            except Exception as _e:
                print(f"[Excel] WARN: validation for {sheet_name!r} failed: {_e}", flush=True)
            finally:
                _sheet_t1 = time.perf_counter()
                try:
                    print(f"[Excel] [{sheet_name}] elapsed={_sheet_t1 - _sheet_t0:.3f}s", flush=True)
                except Exception:
                    pass

        # Routing summary (optional)
        include_summary_cfg = bool(CFG.get("export", {}).get("statements_workbook", {}).get("include_routing_summary", True))
        include_summary_env = str(os.getenv("FRACTO_INCLUDE_ROUTING_SUMMARY", "false")).strip().lower() in ("1","true","yes","y","on")
        if (include_summary_cfg or include_summary_env) and routing_used:
            summary_cols = ["Doc Type", "Parser App ID", "Model", "Extra Accuracy", "Company Type", "Rows"]
            rows = []
            for dt in sheet_order:
                if dt in (routing_used or {}):
                    cfg = routing_used.get(dt, {})
                    try:
                        row_count = int((combined_sheets.get(dt) or {}).shape[0]) if dt in combined_sheets and combined_sheets[dt] is not None else 0
                    except Exception:
                        row_count = 0
                    rows.append([dt, cfg.get("parser_app",""), cfg.get("model",""), str(cfg.get("extra","")), cfg.get("company_type",""), row_count])
            if rows:
                pd.DataFrame(rows, columns=summary_cols).to_excel(writer, sheet_name="Routing Summary", index=False)
                wrote_any_real_sheet = True
                ws_sum = writer.book["Routing Summary"]
                header_font  = Font(bold=True, color=header_font_color)
                header_fill  = PatternFill("solid", fgColor=header_fill_hex)
                header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
                for col_cells in ws_sum.iter_cols(min_row=1, max_row=ws_sum.max_row):
                    longest = max(len("" if c.value is None else str(c.value)) for c in col_cells)
                    ws_sum.column_dimensions[col_cells[0].column_letter].width = min(max(longest + 2, 10), 60)
                    for c in col_cells[1:]:
                        c.alignment = Alignment(vertical="top", wrap_text=True)
                for cell in ws_sum[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                ws_sum.freeze_panes = "A2"

        # Optional "Validation" sheet with sum checks (disabled in base workbook; see validations-only workbook below)
        try:
            if False and INCLUDE_VALIDATION_SHEET and validation_rows:
                import pandas as pd
                from datetime import datetime
                from openpyxl.utils import get_column_letter
                val_cols = [
                    "Doc Type", "Particulars", "Section", "Column", "Sum of Items", "Reported Total", "Difference", "Status", "Tolerance", "Details"
                ]
                pd.DataFrame(validation_rows, columns=val_cols).to_excel(writer, sheet_name=VALIDATION_SHEET_NAME[:31] or "Validation", index=False)
                wrote_any_real_sheet = True
                ws_v = writer.book[VALIDATION_SHEET_NAME[:31] or "Validation"]
                header_font  = Font(bold=True, color=header_font_color)
                header_fill  = PatternFill("solid", fgColor=header_fill_hex)
                header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
                for col_cells in ws_v.iter_cols(min_row=1, max_row=ws_v.max_row):
                    longest = max(len("" if c.value is None else str(c.value)) for c in col_cells)
                    ws_v.column_dimensions[col_cells[0].column_letter].width = min(max(longest + 2, 10), 60)
                for cell in ws_v[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                # Insert a banner row with timestamp and counts
                try:
                    total_checks = len(validation_rows)
                    mismatch_count = sum(1 for r in validation_rows if str(r[7]).strip().upper() == "MISMATCH")
                    doc_types = sorted({str(r[0]) for r in validation_rows})
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    banner = f"Validation generated {ts} — checks: {total_checks}, mismatches: {mismatch_count}, docs: {', '.join(doc_types)}"
                    ws_v.insert_rows(1)
                    last_col = get_column_letter(len(val_cols))
                    ws_v.merge_cells(f"A1:{last_col}1")
                    ws_v["A1"].value = banner
                    ws_v["A1"].font = Font(bold=True)
                    ws_v["A1"].alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                    ws_v["A1"].fill = PatternFill("solid", fgColor="FFF4CC")
                    # Keep header styling on new row 2
                    for cell in ws_v[2]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                    ws_v.freeze_panes = "A3"
                except Exception:
                    ws_v.freeze_panes = "A2"

                # Print concise mismatch summary per doc type
                try:
                    mismatches = [r for r in validation_rows if str(r[7]).upper() == "MISMATCH"]
                    if mismatches:
                        # per doc type counts
                        per_doc: dict[str, int] = {}
                        for r in mismatches:
                            per_doc[str(r[0])] = per_doc.get(str(r[0]), 0) + 1
                        vprint("Validation mismatches by doc:", per_doc)
                        # show top 5 examples
                        for r in mismatches[:5]:
                            vprint("Example:", {"doc": r[0], "particulars": r[1], "section": r[2], "col": r[3], "sum": r[4], "reported": r[5], "diff": r[6], "tol": r[8]})
                    # Always log sheet creation with row count
                    print(f"[Excel] Validation sheet written: rows={len(validation_rows)}", flush=True)
                except Exception:
                    pass
        except Exception as _e:
            print(f"[Excel] WARN: could not write validation sheet: {_e}", flush=True)

        # Optional "Periods" sheet aggregating period metadata for all documents
        try:
            any_periods = any(bool(v) for v in (period_by_doc or {}).values())
        except Exception:
            any_periods = False
        if any_periods:
            period_rows = []
            for dt in sheet_order:
                pdata = (period_by_doc or {}).get(dt, {}) or {}
                for cid, info in pdata.items():
                    period_rows.append([
                        dt,
                        str(cid).upper(),
                        (info or {}).get("label", ""),
                        (info or {}).get("start_date", ""),
                        (info or {}).get("end_date", ""),
                        (info or {}).get("role", ""),
                        "Yes" if (info or {}).get("is_cumulative") else "No",
                        "Yes" if (info or {}).get("is_audited") else "No",
                    ])
            if period_rows:
                pd.DataFrame(
                    period_rows,
                    columns=["Doc Type", "Column ID", "Label", "Start Date", "End Date", "Role", "Cumulative?", "Audited?"]
                ).to_excel(writer, sheet_name="Periods", index=False)
                wrote_any_real_sheet = True
                ws_p = writer.book["Periods"]
                header_font  = Font(bold=True, color=header_font_color)
                header_fill  = PatternFill("solid", fgColor=header_fill_hex)
                header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
                for col_cells in ws_p.iter_cols(min_row=1, max_row=ws_p.max_row):
                    longest = max(len("" if c.value is None else str(c.value)) for c in col_cells)
                    ws_p.column_dimensions[col_cells[0].column_letter].width = min(max(longest + 2, 10), 60)
                    for c in col_cells[1:]:
                        c.alignment = Alignment(vertical="top", wrap_text=True)
                for cell in ws_p[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                ws_p.freeze_panes = "A2"

        # If nothing was written and no seed handling occurred, add a minimal visible sheet so the workbook is valid
        if not wrote_any_real_sheet:
            try:
                import pandas as _pd
                summary_rows = []
                try:
                    for _name, _df in (combined_sheets or {}).items():
                        _rows = int(getattr(_df, 'shape', (0, 0))[0]) if _df is not None else 0
                        summary_rows.append([str(_name), _rows])
                except Exception:
                    pass
                if not summary_rows:
                    summary_rows = [["No visible data", 0]]
                _pd.DataFrame(summary_rows, columns=["Sheet (input)", "Row count"]).to_excel(
                    writer, sheet_name="Summary", index=False
                )
                try:
                    ws_sum = writer.book["Summary"]
                    header_font  = Font(bold=True, color=header_font_color)
                    header_fill  = PatternFill("solid", fgColor=header_fill_hex)
                    header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
                    for cell in ws_sum[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                    ws_sum.freeze_panes = "A2"
                except Exception:
                    pass
            except Exception:
                # Fallback placeholder failed; ignore to avoid aborting save
                pass

    # At the end, optionally write periods debug JSON
    try:
        if debug_flag_from_cfg("IWEALTH_DEBUG_PERIODS_DUMP", "periods_dump", False):
            debug_dump["sheet_order"] = list(sheet_order)
            # quick flag showing if a sheet carried at least one cN column before rename
            for s, meta in debug_dump["sheets"].items():
                cols_before = [str(c).strip().lower() for c in (meta.get("columns_before") or [])]
                meta["had_c_columns_before"] = any(re.fullmatch(r"c\d+", c) for c in cols_before)
            dbg_path = out_path.with_name(f"{out_path.stem}_periods_debug.json")
            with open(dbg_path, "w", encoding="utf-8") as fh:
                json.dump(debug_dump, fh, ensure_ascii=False, indent=2)
            print(f"[Excel] Periods debug written → {dbg_path}")
    except Exception as e:
        print(f"[Excel] Failed to write periods debug: {e}")
    
    try:
        # Verify on-disk workbook details
        _t_verify0 = time.perf_counter()
        wb_verify = load_workbook(out_path)
        _t_verify1 = time.perf_counter()
        _size = os.path.getsize(out_path) if os.path.exists(out_path) else -1
        print(
            f"[Excel] DONE writing workbook: {out_path}\n"
            f"        sheets={wb_verify.sheetnames} size={_size}B load_time={_t_verify1 - _t_verify0:.3f}s",
            flush=True,
        )
        # Robust fallback: if too few sheets were persisted, rewrite in simple mode
        try:
            expected = sum(1 for _name, _df in (combined_sheets or {}).items() if _df is not None and not getattr(_df, 'empty', True))
            actual   = len([s for s in wb_verify.sheetnames])
            if expected >= 2 and actual <= 1:
                print("[Excel] WARN: workbook saved with fewer sheets than expected — rewriting in simple mode…", flush=True)
                import pandas as _pd
                with _pd.ExcelWriter(out_path, engine="openpyxl") as _w2:
                    # 1) Statement sheets (simple write, ordered)
                    for sname in sheet_order:
                        df0 = combined_sheets.get(sname)
                        if df0 is None or getattr(df0, 'empty', True):
                            continue
                        try:
                            df2 = _normalize_df_for_excel(sname, df0.copy())
                        except Exception:
                            df2 = df0
                        df2 = _sanitize_df_for_excel(df2)
                        _safe = (sname[:31] or 'Sheet')
                        df2.to_excel(_w2, sheet_name=_safe, index=False)

                    # 2) Validation sheet (optional, de-duplicated)
                    try:
                        if INCLUDE_VALIDATION_SHEET and validation_rows:
                            val_cols = [
                                "Doc Type", "Particulars", "Section", "Column", "Sum of Items", "Reported Total", "Difference", "Status", "Tolerance", "Details"
                            ]
                            (_pd.DataFrame(validation_rows, columns=val_cols)
                                .drop_duplicates()
                                .to_excel(_w2, sheet_name=(VALIDATION_SHEET_NAME[:31] or "Validation"), index=False))
                    except Exception:
                        pass

                    # 3) Routing Summary (write minimal summary if routing_used missing)
                    try:
                        include_summary_cfg = bool(CFG.get("export", {}).get("statements_workbook", {}).get("include_routing_summary", True))
                        include_summary_env = str(os.getenv("FRACTO_INCLUDE_ROUTING_SUMMARY", "false")).strip().lower() in ("1","true","yes","y","on")
                        if include_summary_cfg or include_summary_env:
                            rows = []
                            if routing_used:
                                for dt in sheet_order:
                                    if dt in (routing_used or {}):
                                        cfg = routing_used.get(dt, {})
                                        try:
                                            row_count = int((combined_sheets.get(dt) or {}).shape[0]) if dt in combined_sheets and combined_sheets[dt] is not None else 0
                                        except Exception:
                                            row_count = 0
                                        rows.append([dt, cfg.get("parser_app",""), cfg.get("model",""), str(cfg.get("extra","")), row_count])
                                cols = ["Doc Type", "Parser App ID", "Model", "Extra Accuracy", "Rows"]
                            else:
                                for dt in sheet_order:
                                    try:
                                        row_count = int((combined_sheets.get(dt) or {}).shape[0]) if dt in combined_sheets and combined_sheets[dt] is not None else 0
                                    except Exception:
                                        row_count = 0
                                    rows.append([dt, row_count])
                                cols = ["Doc Type", "Rows"]
                            if rows:
                                _pd.DataFrame(rows, columns=cols).to_excel(_w2, sheet_name="Routing Summary", index=False)
                    except Exception:
                        pass

                    # 4) Periods sheet (optional)
                    try:
                        any_periods = any(bool(v) for v in (period_by_doc or {}).values())
                    except Exception:
                        any_periods = False
                    try:
                        if any_periods:
                            period_rows = []
                            for dt in sheet_order:
                                pdata = (period_by_doc or {}).get(dt, {}) or {}
                                for cid, info in pdata.items():
                                    period_rows.append([
                                        dt,
                                        str(cid).upper(),
                                        (info or {}).get("label", ""),
                                        (info or {}).get("start_date", ""),
                                        (info or {}).get("end_date", ""),
                                        (info or {}).get("role", ""),
                                        "Yes" if (info or {}).get("is_cumulative") else "No",
                                        "Yes" if (info or {}).get("is_audited") else "No",
                                    ])
                            if period_rows:
                                _pd.DataFrame(period_rows, columns=["Doc Type", "Column ID", "Label", "Start Date", "End Date", "Role", "Cumulative?", "Audited?"]).to_excel(_w2, sheet_name="Periods", index=False)
                    except Exception:
                        pass
                # Re-open and report
                wb_verify = load_workbook(out_path)
                print(f"[Excel] Rewrite complete — sheets now: {wb_verify.sheetnames}", flush=True)
        except Exception as _re:
            print(f"[Excel] WARN: simple rewrite attempt failed: {_re}", flush=True)
    except Exception as _ve:
        try:
            _size = os.path.getsize(out_path) if os.path.exists(out_path) else -1
            print(f"[Excel] DONE writing workbook: {out_path} (size={_size}B). Verify failed: {_ve}", flush=True)
        except Exception:
            print(f"[Excel] DONE writing workbook: {out_path}. Verify failed: {_ve}", flush=True)

    try:
        t1 = time.perf_counter()
        print(f"[Excel] Total elapsed={t1 - t0:.3f}s", flush=True)
    except Exception:
        pass

    # Also write a client-friendly workbook: drop meta cols, keep non-empty statements + Validation, exclude Periods
    try:
        client_cfg = (CFG.get("export", {}) or {}).get("filenames", {}) or {}
        client_name = client_cfg.get("client_xlsx", "{stem}_client.xlsx")
        client_path = Path(pdf_path).expanduser().resolve().with_name(client_name.format(stem=stem))
        import pandas as pd
        with pd.ExcelWriter(client_path, engine="openpyxl") as wc:
            try:
                print(f"[Excel][client] available period docs: {list((period_labels_by_doc or {}).keys())}", flush=True)
            except Exception:
                pass
            def _prepare_client_df(name: str, df_in: "pd.DataFrame"):
                import re as _re
                if df_in is None or getattr(df_in, "empty", True):
                    return None
                df = _normalize_df_for_excel(name, df_in.copy())

                # Rename period headers if configured
                def _use_labels_for_sheet_local(name_: str) -> bool:
                    val = use_period_labels_cfg
                    if isinstance(val, dict):
                        if name_ in val:
                            return bool(val[name_])
                        for k, v in val.items():
                            try:
                                if str(k).strip().lower() == str(name_).strip().lower():
                                    return bool(v)
                            except Exception:
                                continue
                        return bool(val.get("default", True))
                    return bool(val)
                if _use_labels_for_sheet_local(name):
                    try:
                        period_labels = _pick_period_labels_for_sheet(name, period_labels_by_doc, PERIOD_LABELS_BY_DOC)
                        if period_labels and any(isinstance(v, dict) for v in period_labels.values()):
                            period_labels = {str(k).lower(): (v.get("label") or "") if isinstance(v, dict) else str(v) for k, v in period_labels.items()}
                        else:
                            period_labels = {str(k).lower(): ("" if v is None else str(v)) for k, v in (period_labels or {}).items()}
                        rename_map: dict[str, str] = {}
                        has_indexed_keys = any(_re.fullmatch(r"(?i)[cp]\\d+", str(k)) for k in period_labels.keys()) if period_labels else False
                        if has_indexed_keys:
                            for col in list(df.columns):
                                low = str(col).strip().lower()
                                m = _re.fullmatch(r"([cp])(\\d+)", low)
                                if not m:
                                    continue
                                prefix, num = m.group(1), m.group(2)
                                key = f"{prefix}{num}"
                                alt = f"{'p' if prefix == 'c' else 'c'}{num}"
                                label = (period_labels or {}).get(key) or (period_labels or {}).get(alt)
                                if label:
                                    rename_map[col] = str(label)
                        else:
                            try:
                                c_cols = sorted(
                                    [c for c in df.columns if _re.fullmatch(r"(?i)c\\d+", str(c))],
                                    key=lambda x: int(_re.findall(r"\\d+", str(x))[0])
                                )
                            except Exception:
                                c_cols = [c for c in df.columns if str(c).lower().startswith('c')]
                            labels_in_order = [str(v) for v in (period_labels or {}).values() if str(v).strip() != ""]
                            n = min(len(c_cols), len(labels_in_order))
                            for i in range(n):
                                rename_map[c_cols[i]] = labels_in_order[i]
                        print(f"[Excel][client] period-label-keys for {name!r}: {list((period_labels or {}).keys())}", flush=True)
                        print(f"[Excel][client] rename-map for {name!r}: {rename_map}", flush=True)
                        if rename_map:
                            df = df.rename(columns=rename_map)
                        elif isinstance(debug_dump.get('sheets'), dict) and name in debug_dump['sheets']:
                            # Reuse debug rename_map if available
                            try:
                                rm = (debug_dump['sheets'].get(name) or {}).get('rename_map') or {}
                                if rm:
                                    print(f"[Excel][client] using debug rename-map for {name!r}: {rm}", flush=True)
                                    df = df.rename(columns=rm)
                            except Exception:
                                pass
                    except Exception:
                        pass

                # Drop requested meta columns
                drop_meta = {"id","row_type","parent_id","section","calculation_references"}
                cols_to_drop = [c for c in df.columns if str(c).strip().lower() in drop_meta]
                if cols_to_drop:
                    df = df.drop(columns=cols_to_drop)

                # Drop empty numeric columns
                meta_keep = {"particulars","sr_no"}
                empties = []
                for c in list(df.columns):
                    cl = str(c).strip().lower()
                    if cl in meta_keep:
                        continue
                    coerced = df[c].apply(_coerce_number_like)
                    if sum(v is not None for v in coerced) == 0:
                        empties.append(c)
                if empties:
                    df = df.drop(columns=empties)

                # Keep only non-empty sheets with at least one period column
                non_meta_cols = [c for c in df.columns if str(c).strip().lower() not in meta_keep]
                if df.empty or len(non_meta_cols) == 0:
                    return None
                # at least one numeric present
                has_num = any(df[c].apply(_coerce_number_like).notna().any() for c in non_meta_cols)
                if not has_num:
                    return None
                return df

            header_font  = Font(bold=True, color=header_font_color)
            header_fill  = PatternFill("solid", fgColor=header_fill_hex)
            header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)

            written_any = False
            for sheet_name in sheet_order:
                # Only Cashflow, Balance Sheet, and Profit & Loss
                name_l = str(sheet_name).strip().lower()
                if not (("cash" in name_l and "flow" in name_l) or ("balance" in name_l and "sheet" in name_l) or ("profit" in name_l and "loss" in name_l)):
                    continue
                df0 = combined_sheets.get(sheet_name)
                dfc = _prepare_client_df(sheet_name, df0)
                if dfc is None:
                    continue
                safe = sheet_name[:31] or "Sheet"
                dfc.to_excel(wc, sheet_name=safe, index=False)
                ws = wc.book[safe]
                for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
                    longest = max(len("" if c.value is None else str(c.value)) for c in col_cells)
                    ws.column_dimensions[col_cells[0].column_letter].width = min(max(longest + 2, 10), 60)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                ws.freeze_panes = "A2"
                written_any = True

            # Validation sheet (optional)
            if INCLUDE_VALIDATION_SHEET and validation_rows:
                val_cols = [
                    "Doc Type", "Particulars", "Section", "Column", "Sum of Items", "Reported Total", "Difference", "Status", "Tolerance", "Details"
                ]
                # De-duplicate identical validation rows to avoid repetition in the sheet
                val_df = pd.DataFrame(validation_rows, columns=val_cols).drop_duplicates()
                vname = VALIDATION_SHEET_NAME[:31] or "Validation"
                val_df.to_excel(wc, sheet_name=vname, index=False)
                ws_v = wc.book[vname]
                for col_cells in ws_v.iter_cols(min_row=1, max_row=ws_v.max_row):
                    longest = max(len("" if c.value is None else str(c.value)) for c in col_cells)
                    ws_v.column_dimensions[col_cells[0].column_letter].width = min(max(longest + 2, 10), 60)
                for cell in ws_v[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                ws_v.freeze_panes = "A2"
                written_any = True

            # Ensure at least one visible sheet exists (fallback Summary)
            if not written_any:
                import pandas as _pd
                _pd.DataFrame([["No visible data"]], columns=["Status"]).to_excel(wc, sheet_name="Summary", index=False)
                ws_s = wc.book["Summary"]
                for cell in ws_s[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                ws_s.freeze_panes = "A2"

        print(f"[Excel] Client workbook written → {client_path}", flush=True)
    except Exception as e:
        print(f"[Excel] WARN: failed to write client workbook: {e}", flush=True)

    # Also write a separate validations-only workbook for easy debugging
    try:
        if INCLUDE_VALIDATION_SHEET:
            import pandas as pd
            from datetime import datetime
            from openpyxl.utils import get_column_letter
            val_only_path = Path(pdf_path).expanduser().resolve().with_name(f"{stem}_validations.xlsx")
            val_cols = [
                "Doc Type", "Particulars", "Section", "Column", "Sum of Items", "Reported Total", "Difference", "Status", "Tolerance", "Details"
            ]
            with pd.ExcelWriter(val_only_path, engine="openpyxl") as wv:
                # Always produce the sheet, even if empty
                # De-duplicate to remove repeated validations
                val_df = pd.DataFrame(validation_rows, columns=val_cols).drop_duplicates()
                sheet_name = VALIDATION_SHEET_NAME[:31] or "Validation"
                val_df.to_excel(wv, sheet_name=sheet_name, index=False)
                ws = wv.book[sheet_name]
                header_font  = Font(bold=True, color=header_font_color)
                header_fill  = PatternFill("solid", fgColor=header_fill_hex)
                header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
                # Autosize columns
                for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
                    longest = max(len("" if c.value is None else str(c.value)) for c in col_cells)
                    ws.column_dimensions[col_cells[0].column_letter].width = min(max(longest + 2, 10), 60)
                # Style header row
                if ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                # Banner row
                try:
                    # Summary banner based on de-duplicated validations
                    total_checks = int(len(val_df))
                    mismatch_count = int((val_df["Status"].astype(str).str.upper() == "MISMATCH").sum()) if not val_df.empty else 0
                    doc_types = sorted(val_df["Doc Type"].astype(str).unique().tolist()) if not val_df.empty else []
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    banner = f"Validation generated {ts} — checks: {total_checks}, mismatches: {mismatch_count}, docs: {', '.join(doc_types) if doc_types else '-'}"
                    ws.insert_rows(1)
                    last_col = get_column_letter(len(val_cols))
                    ws.merge_cells(f"A1:{last_col}1")
                    ws["A1"].value = banner
                    ws["A1"].font = Font(bold=True)
                    ws["A1"].alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                    ws["A1"].fill = PatternFill("solid", fgColor="FFF4CC")
                    # Re-apply header style on new header row (row 2)
                    if ws.max_row >= 2:
                        for cell in ws[2]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_align
                    ws.freeze_panes = "A3"
                except Exception:
                    try:
                        ws.freeze_panes = "A2"
                    except Exception:
                        pass
            print(f"[Excel] Validation-only workbook written → {val_only_path} (rows={len(val_df)})", flush=True)
    except Exception as e:
        print(f"[Excel] WARN: failed to write validations-only workbook: {e}", flush=True)
    return str(out_path)

def write_excel_from_ocr(
    results: List[Dict[str, Any]],
    output_path: str | io.BytesIO,
    overrides: dict[str, str] | None = None,
    *,
    mappings: dict[str, str] | None = None,
    template_path: str | None = None,
    sheet_name: str | None = None,
):
    """
    Write OCR rows to *output_path*.

    Parameters
    ----------
    results : list[dict]
        Fracto API responses (or pre‑loaded JSON) – list of results.
    output_path : str | io.BytesIO
        Where to write the Excel workbook.
    overrides : dict[str, str], optional
        {column_name: value} pairs forced into every row (e.g. constant HS‑Code).
    mappings : dict[str, str], optional
        Column → source‑field mapping. Defaults to the global MAPPINGS.
    template_path : str | Path, optional
        Path to an .xlsx template to use as a base (preserves styles).
    sheet_name : str, optional
        Which sheet inside the template/workbook to write into. Defaults to the
        first/active sheet.
    """
    mappings = mappings or MAPPINGS
    template_path = template_path or TEMPLATE_PATH
    sheet_name = sheet_name or SHEET_NAME
    overrides = overrides or {}

    headers = list(mappings.keys())

    # Keep only overrides whose column exists in the header list
    overrides = {k: v for k, v in overrides.items() if k in headers}

    # Load or create workbook
    if template_path and Path(template_path).expanduser().exists():
        wb = load_workbook(Path(template_path).expanduser())
    else:
        wb = Workbook()

    # Select or create target sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Clear existing data
    ws.delete_rows(1, ws.max_row)

    # ── Gather rows from results ──
    all_rows: list[dict] = []
    for result in results:
        payload = result.get("data", [])
        rows = _extract_rows(payload)
        all_rows.extend(rows)

    # Header row
    ws.append(headers)

    # Write data rows
    written = 0
    for row in all_rows:
        excel_row = []
        for col in headers:
            src_field = mappings.get(col, col)
            value = overrides.get(col, row.get(src_field, ""))
            excel_row.append(value)
        ws.append(excel_row)
        written += 1

    # Convert numeric-like columns to proper numbers
    import pandas as pd
    for idx, _ in enumerate(headers, start=1):
        values = [ws.cell(row=r, column=idx).value for r in range(2, ws.max_row + 1)]
        series = pd.Series(values).replace("", pd.NA)
        converted = pd.to_numeric(series, errors="coerce")
        if converted.isna().eq(series.isna()).all():
            for r, val in enumerate(converted, start=2):
                if pd.isna(val):
                    ws.cell(row=r, column=idx).value = None
                else:
                    fval = float(val)
                    ws.cell(row=r, column=idx).value = int(fval) if fval.is_integer() else fval

    # ── Styling (same as before) ──
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="305496")
    header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
    thin_border  = Side(border_style="thin", color="999999")
    border       = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    max_width = 60
    for column in ws.iter_cols(min_row=1, max_row=ws.max_row):
        longest = max(len(str(c.value)) if c.value is not None else 0 for c in column)
        width = min(max(longest + 2, 10), max_width)
        ws.column_dimensions[column[0].column_letter].width = width
        for c in column[1:]:
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    # Zebra striping for readability
    stripe_fill = PatternFill("solid", fgColor="F2F2F2")
    for r in range(2, ws.max_row + 1):
        if r % 2 == 0:
            for c in ws[r]:
                c.fill = stripe_fill

    # Save
    if isinstance(output_path, io.BytesIO):
        wb.save(output_path)
    else:
        wb.save(str(output_path))

    logger.info(
        "Excel written to %s (%d rows, %d columns)",
        output_path if isinstance(output_path, str) else "<buffer>",
        written,
        len(headers),
    )


def _write_statements_json(
    pdf_path: str,
    stem: str,
    combined_rows: dict[str, list[dict]],
    groups: dict[str, list[int]] | None,
    routing_used: dict[str, dict] | None,
    company_type: str | None,
    out_path_override: str | None = None,
    first_pass_results: list[dict] | None = None,
    second_pass_result: dict | None = None,
    third_pass_raw: dict[str, list[dict]] | None = None,
) -> str:
    """
    Write combined statements JSON with rows, pages, routing, and period metadata.
    """
    # 1) Periods from in-memory third-pass results
    periods_by_doctype: dict[str, dict] = {}
    _labels_for_excel: dict[str, dict] = {}
    try:
        if third_pass_raw:
            for _dt_key, _res_list in (third_pass_raw or {}).items():
                dt_norm = normalize_doc_type(_dt_key)
                candidates = _res_list if isinstance(_res_list, list) else [_res_list]
                for _res in candidates:
                    if not isinstance(_res, dict):
                        continue
                    pd_payload = (((_res.get("data") or {}).get("parsedData")) or {})
                    by_id, labels = _extract_period_maps_from_payload(pd_payload)
                    if by_id:
                        periods_by_doctype[dt_norm] = by_id
                        _labels_for_excel[dt_norm] = {k.lower(): v for k, v in labels.items()}
                        break
        if _labels_for_excel:
            for _k, _v in _labels_for_excel.items():
                PERIOD_LABELS_BY_DOC[_k] = _v
    except Exception:
        pass

    # 2) Fallback: scan disk for *_ocr.json per-group files
    try:
        _by_doc, _labels = _scan_group_jsons_for_periods(pdf_path, stem)
        if _by_doc:
            for _k, _v in _by_doc.items():
                periods_by_doctype.setdefault(_k, {}).update(_v)
        if _labels:
            for _k, _v in _labels.items():
                PERIOD_LABELS_BY_DOC.setdefault(_k, {}).update(_v)
    except Exception:
        pass

    allowed = [lbl for lbl in (CFG.get("labels", {}).get("canonical", []) or []) if lbl != "Others"]

    def _coerce_row_numbers(row: dict) -> dict:
        out = dict(row)
        try:
            import re as _re
            for k in list(out.keys()):
                if _re.fullmatch(r"(?i)[cp]\d+", str(k)):
                    v = out[k]
                    nv = _coerce_number_like(v)
                    if v in ("", None):
                        continue
                    out[k] = nv if nv is not None else out[k]
        except Exception:
            pass
        return out

    docs: dict[str, dict] = {}
    for doc_type in allowed:
        rows = combined_rows.get(doc_type) or []
        try:
            rows = [_coerce_row_numbers(r) for r in rows]
        except Exception:
            pass
        if not rows:
            continue
        meta = (routing_used or {}).get(doc_type, {}) if routing_used else {}
        page_list = (groups or {}).get(doc_type, []) if groups else []
        entry = {
            "pages": page_list,
            "parser_app": meta.get("parser_app", ""),
            "model": meta.get("model", ""),
            "extra_accuracy": meta.get("extra", ""),
            "periods": periods_by_doctype.get(doc_type, {}),
        }
        include_rows = bool((CFG.get("export", {}).get("combined_json", {}) or {}).get("include_rows", True))
        if include_rows:
            entry["rows"] = rows
        docs[doc_type] = entry

    out = {
        "file": Path(pdf_path).name,
        "status": "ok",
        "company_type": company_type or "",
        "documents": docs,
    }
    try:
        _dbg_docs = {k: sorted((v or {}).get("periods", {}).keys()) for k, v in (docs or {}).items()}
        logger.info("Combined JSON: periods per doc → %s", _dbg_docs)
    except Exception:
        pass

    # Optionally include extra sections
    combined_json_cfg = (CFG.get("export", {}).get("combined_json", {}) or {})
    if combined_json_cfg.get("include_first_pass") and first_pass_results is not None:
        out["first_pass"] = first_pass_results
    if combined_json_cfg.get("include_second_pass") and second_pass_result is not None:
        out["second_pass"] = second_pass_result
    if combined_json_cfg.get("include_third_pass_raw") and third_pass_raw:
        out["third_pass"] = third_pass_raw

    # Output path
    json_name_tmpl = combined_json_cfg.get("filename") \
        or CFG.get("export", {}).get("filenames", {}).get("statements_json") \
        or "{stem}_statements.json"
    out_path = Path(out_path_override).expanduser().resolve() if out_path_override else Path(pdf_path).expanduser().resolve().with_name(json_name_tmpl.format(stem=stem))
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(out, fh, indent=2)
    logger.info("Combined JSON written to %s", out_path)
    return str(out_path)

def generate_statements_excel(pdf_bytes: bytes, original_filename: str) -> bytes | None:
    """
    Robust multi-sheet workbook creator:
      • 1st pass (per-page) to shortlist pages
      • Expand selection by ±1 neighbour page to catch 'continued' pages
      • 2nd pass to classify
      • Header-based heuristics + smoothing to fix obviously wrong labels / 'Others'
      • 3rd pass per doc_type; each statement in its own sheet
    Returns workbook bytes or None.
    """
    # 1) First pass (optional)
    first_enabled = bool(CFG.get("passes", {}).get("first", {}).get("enable", True))
    results = []
    stem = Path(original_filename).stem

    if first_enabled:
        results = call_fracto_parallel(pdf_bytes, original_filename, extra_accuracy=EXTRA_ACCURACY_FIRST)
        total_pages = len(results) if results else 0
        selected_pages = [
            idx + 1
            for idx, res in enumerate(results or [])
            if (res.get("data", {}).get("parsedData", {}).get("Document_type", "Others") or "Others").strip().lower() != "others"
        ]
        selected_pages = expand_selected_pages(selected_pages, total_pages, radius=1)
    else:
        total_pages = get_page_count_from_bytes(pdf_bytes)
        selected_pages = list(range(1, total_pages + 1))

    if not selected_pages:
        return None

    filtered_pages, mid_diag = filter_pages_via_mid_pass(
        pdf_bytes,
        selected_pages,
        stem=stem,
        logger_obj=logger,
    )
    if mid_diag:
        try:
            summary = "; ".join(
                f"{item['page']}={item['label'] or 'unclassified'}{'(drop)' if not item['keep'] else ''}"
                for item in mid_diag
            )
            logger.info("[mid-pass] page classifications → %s", summary)
        except Exception:
            pass
        dropped = [item["page"] for item in mid_diag if not item.get("keep")]
        if dropped:
            logger.info("[mid-pass] dropping pages before second pass: %s", dropped)
    if filtered_pages:
        selected_pages = filtered_pages
    elif mid_diag:
        logger.warning("[mid-pass] classifier removed all pages; reverting to previous selection")

    # Build selected.pdf
    from iwe_core.pdf_ops import build_pdf_from_pages
    selected_bytes = build_pdf_from_pages(pdf_bytes, selected_pages)

    # 2) Second pass
    second_res = call_fracto(
        selected_bytes,
        f"{stem}_selected.pdf",
        parser_app=SECOND_PARSER_APP_ID,
        model=SECOND_MODEL_ID,
        extra_accuracy=EXTRA_ACCURACY_SECOND,
    )

    # 3) Classification (with fallback)
    classification = (
        second_res.get("data", {}).get("parsedData", {}).get("page_wise_classification", [])
        if isinstance(second_res, dict) else []
    ) or []
    if not classification:
        tmp: list[dict] = []
        if results:
            for sel_idx, orig_page in enumerate(selected_pages, start=1):
                if not (1 <= orig_page <= len(results)):
                    continue
                res = results[orig_page - 1] or {}
                dt = (res.get("data", {}) or {}).get("parsedData", {}).get("Document_type")
                if dt and str(dt).strip().lower() != "others":
                    tmp.append({"page_number": sel_idx, "doc_type": dt})
        elif mid_diag:
            label_map = {int(it.get("page", 0)): str(it.get("label") or "") for it in mid_diag if isinstance(it, dict)}
            for sel_idx, orig_page in enumerate(selected_pages, start=1):
                raw = label_map.get(orig_page, "")
                if not raw:
                    continue
                human = raw.replace("_", " ")
                dt_norm = normalize_doc_type(human)
                if dt_norm and dt_norm != "Others":
                    tmp.append({"page_number": sel_idx, "doc_type": dt_norm})
        classification = tmp
    groups = build_groups(selected_pages, classification, pdf_bytes)
    if not groups:
        return None

    # 4) Third pass per group (parallelized)
    import pandas as pd
    combined_sheets: dict[str, pd.DataFrame] = {}
    routing_used: dict[str, dict] = {}

    work_items: list[tuple[str, list[int], tuple[str, str, str]]] = []
    for doc_type, page_list in groups.items():
        page_list = sorted(page_list)
        parser_app, model_id, extra_acc = _resolve_routing(doc_type)
        routing_used[doc_type] = {"parser_app": parser_app, "model": model_id, "extra": extra_acc}
        work_items.append((doc_type, page_list, (parser_app, model_id, extra_acc)))

    def _process_group(dt: str, pages: list[int], parser_app: str, model_id: str, extra_acc: str):
        try:
            gb = build_pdf_from_pages(pdf_bytes, pages)
            res = call_fracto(
                gb,
                f"{stem}_{dt.lower().replace(' ', '_').replace('&','and').replace('/','_')}.pdf",
                parser_app=parser_app,
                model=model_id,
                extra_accuracy=extra_acc,
            )
            parsed = res.get("data", {}).get("parsedData", [])
            df = None
            if isinstance(parsed, list) and parsed:
                all_keys = []
                for row in parsed:
                    for k in row.keys():
                        if k not in all_keys:
                            all_keys.append(k)
                rows = [{k: r.get(k, "") for k in all_keys} for r in parsed]
                _pd = __import__('pandas')
                df = _pd.DataFrame(rows, columns=all_keys)
                df = sanitize_statement_df(dt, df)
                try:
                    import os as _os
                    if str(_os.getenv("IWEALTH_ENABLE_REORDER", "0")).strip() in {"1", "true", "yes"}:
                        df = reorder_dataframe_sections_first(df)
                except Exception:
                    pass
            return (dt, df)
        except Exception:
            return (dt, None)

    max_workers = int((CFG.get("concurrency", {}) or {}).get("max_parallel", 9)) or 4
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = [pool.submit(_process_group, dt, pages, pa, mid, ex) for (dt, pages, (pa, mid, ex)) in work_items]
        for fut in as_completed(futures):
            dt, df = fut.result()
            if df is not None:
                combined_sheets[dt] = df

    if not combined_sheets:
        return None

    # 5) Write workbook to bytes (styled)
    out_buf = io.BytesIO()
    _t_mem0 = time.perf_counter()
    print("[Excel] ENTER in-memory writer (BytesIO), engine=openpyxl", flush=True)
    with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
        try:
            print(f"[Excel] In-memory writer created: book_type={type(writer.book).__name__}", flush=True)
        except Exception:
            pass

        # Collect optional validation rows (Cashflow support; can be extended)
        validation_rows: list[list] = []

        def _cf_validations(sheet_label: str, df_in: "pd.DataFrame") -> None:
            vcfg = (CFG.get("validation", {}) or {}).get("checks", {}) or {}
            cfc = (vcfg.get("cashflow") or {})
            enable = bool(cfc.get("enable", True)) and bool(cfc.get("enforce_cash_tie", True))
            if not enable or df_in is None or getattr(df_in, "empty", True):
                return
            import re as _re
            df = df_in.copy()
            part_col = next((c for c in df.columns if str(c).strip().lower() in {"particulars","particular","description","line item","line_item"}), None)
            if part_col is None:
                return
            tol_pct = float(cfc.get("tolerance_pct", 0.005))
            abs_min = float(cfc.get("abs_min", 1.0))
            # pick numeric columns
            num_cols = [c for c in df.columns if _re.fullmatch(r"(?i)c\d+", str(c))]
            if not num_cols:
                meta_cols = {part_col, "sr_no", "id", "row_type", "parent_id", "components", "calculation_references", "section"}
                num_cols = [c for c in df.columns if str(c) not in meta_cols]
            for c in list(num_cols):
                coerced = df[c].apply(_coerce_number_like)
                if sum(v is not None for v in coerced) == 0:
                    num_cols.remove(c)
                    continue
                df[c] = coerced
            if not num_cols:
                return

            # use module-level _compile and _find_rows_matching

            labels = df[part_col].astype(str)
            op_pats = _compile([
                r"^\s*net\s+cash(?:\s+flow)?\b.*operating\s+activities",
                r"^\s*net\s+cash\s+generated.*operating\s+activities",
                r"^\s*net\s+cash\s+from.*operating\s+activities",
                r"^\s*net\s+cash\s+used.*operating\s+activities",
            ])
            inv_pats = _compile([
                r"^\s*net\s+cash(?:\s+flow)?\b.*investing\s+activities",
                r"^\s*net\s+cash\s+used.*investing\s+activities",
                r"^\s*net\s+cash\s+from.*investing\s+activities",
            ])
            fin_pats = _compile([
                r"^\s*net\s+cash(?:\s+flow)?\b.*financing\s+activities",
                r"^\s*net\s+cash\s+used.*financing\s+activities",
                r"^\s*net\s+cash\s+from.*financing\s+activities",
            ])
            net_pats = _compile([
                r"net\s+(?:increase|decrease).*cash\s*(?:and|&)\s*cash\s*equivalents",
                r"net\s+change\s+in\s+cash\s*(?:and|&)\s*cash\s*equivalents",
            ])
            open_pats = _compile([
                r"cash\s*(?:and|&)\s*cash\s*equivalents.*beginning",
                r"opening\s+cash.*equivalents",
            ])
            close_pats = _compile([
                r"cash\s*(?:and|&)\s*cash\s*equivalents.*end",
                r"closing\s+cash.*equivalents",
            ])

            op_rows = _find_rows_matching(labels, op_pats)
            inv_rows = _find_rows_matching(labels, inv_pats)
            fin_rows = _find_rows_matching(labels, fin_pats)
            net_rows = _find_rows_matching(labels, net_pats)
            open_rows = _find_rows_matching(labels, open_pats)
            close_rows = _find_rows_matching(labels, close_pats)

            # use module-level first-value helper

            for c in num_cols:
                op = _first_val_from_rows(df, op_rows, c)
                inv = _first_val_from_rows(df, inv_rows, c)
                fin = _first_val_from_rows(df, fin_rows, c)
                net = _first_val_from_rows(df, net_rows, c)
                opening = _first_val_from_rows(df, open_rows, c)
                closing = _first_val_from_rows(df, close_rows, c)

                if op is not None and inv is not None and fin is not None and net is not None:
                    total = float(op) + float(inv) + float(fin)
                    diff = total - float(net)
                    tol = max(abs_min, abs(float(net)) * tol_pct)
                    ok = abs(diff) <= tol
                    details = f"CFO={op}; CFI={inv}; CFF={fin}; Net={net}"
                    validation_rows.append([sheet_label, "Cashflow Tie", "CFO+CFI+CFF", str(c), float(total), float(net), diff, "OK" if ok else "MISMATCH", tol, details])

                net2 = net
                if net2 is None and (op is not None and inv is not None and fin is not None):
                    net2 = float(op) + float(inv) + float(fin)
                if opening is not None and closing is not None and net2 is not None:
                    est_close = float(opening) + float(net2)
                    diff = est_close - float(closing)
                    tol = max(abs_min, max(abs(float(est_close)), abs(float(closing))) * tol_pct)
                    ok = abs(diff) <= tol
                    details = f"Opening={opening}; Net={net2}; Closing={closing}"
                    validation_rows.append([sheet_label, "Cash Balance Tie", "Opening→Closing", str(c), float(est_close), float(closing), diff, "OK" if ok else "MISMATCH", tol, details])

        for sheet_name, df in combined_sheets.items():
            _s0 = time.perf_counter()
            safe = sheet_name[:31] or "Sheet"
            try:
                _shape = getattr(df, 'shape', (0,0))
                _mem = int(df.memory_usage(index=True, deep=True).sum()) if hasattr(df, 'memory_usage') else -1
                print(f"[Excel] [mem:{sheet_name}] rows={_shape[0]} cols={_shape[1]} approx_mem={_mem}B -> {safe}", flush=True)
            except Exception:
                pass
            df.to_excel(writer, sheet_name=safe, index=False)
            ws = writer.book[safe]
            header_font  = Font(bold=True, color="FFFFFF")
            header_fill  = PatternFill("solid", fgColor="305496")
            header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
            max_width = 60
            for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
                longest = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                width = min(max(longest + 2, 10), max_width)
                ws.column_dimensions[col[0].column_letter].width = width
                for c in col[1:]:
                    c.alignment = Alignment(vertical="top", wrap_text=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            ws.freeze_panes = "A2"
            try:
                print(f"[Excel] [mem:{sheet_name}] worksheet dims rows={ws.max_row} cols={ws.max_column}", flush=True)
            except Exception:
                pass
            _s1 = time.perf_counter()
            try:
                print(f"[Excel] [mem:{sheet_name}] elapsed={_s1 - _s0:.3f}s", flush=True)
            except Exception:
                pass

            # Per-sheet validations (currently: Cashflow)
            try:
                if INCLUDE_VALIDATION_SHEET and isinstance(df, pd.DataFrame) and not df.empty:
                    name_l = str(sheet_name).strip().lower()
                    is_cf = ("cash" in name_l and "flow" in name_l)
                    if is_cf:
                        _cf_validations(sheet_name, df)
            except Exception:
                pass

        # Optional routing summary
        include_summary = str(os.getenv("FRACTO_INCLUDE_ROUTING_SUMMARY", "false")).strip().lower() in ("1","true","yes","y","on")
        if include_summary and routing_used:
            rows = []
            for dt in sorted(routing_used):
                cfg = routing_used[dt]
                rows.append([dt, cfg.get("parser_app",""), cfg.get("model",""), str(cfg.get("extra",""))])
            pd.DataFrame(rows, columns=["Doc Type","Parser App ID","Model","Extra Accuracy"]).to_excel(writer, sheet_name="Routing Summary", index=False)
            ws = writer.book["Routing Summary"]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="305496")
            wrote_any_real_sheet = True

        # Write Validation sheet if we collected any
        if INCLUDE_VALIDATION_SHEET and validation_rows:
            try:
                import pandas as _pd
                val_cols = [
                    "Doc Type","Particulars","Section","Column","Sum of Items","Reported Total","Difference","Status","Tolerance","Details"
                ]
                # De-duplicate to prevent repeated entries
                _pd.DataFrame(validation_rows, columns=val_cols).drop_duplicates().to_excel(
                    writer, sheet_name=(VALIDATION_SHEET_NAME[:31] or "Validation"), index=False
                )
            except Exception as _e:
                print(f"[Excel] WARN: in-memory validation sheet failed: {_e}", flush=True)
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            ws.freeze_panes = "A2"
            try:
                print(f"[Excel] [mem:Routing Summary] dims rows={ws.max_row} cols={ws.max_column}", flush=True)
            except Exception:
                pass
            wrote_any_real_sheet = True

        # Fallback: if nothing was written (e.g., all sheets filtered as empty/non-numeric),
        # add a minimal visible sheet so the workbook remains valid and debuggable.
        try:
            if not wrote_any_real_sheet:
                import pandas as _pd
                summary_rows = []
                try:
                    for _name, _df in (combined_sheets or {}).items():
                        _rows = int(getattr(_df, 'shape', (0, 0))[0]) if _df is not None else 0
                        summary_rows.append([str(_name), _rows])
                except Exception:
                    pass
                if not summary_rows:
                    summary_rows = [["No visible data", 0]]
                _pd.DataFrame(summary_rows, columns=["Sheet (input)", "Row count"]).to_excel(
                    writer, sheet_name="Summary", index=False
                )
                try:
                    ws_sum = writer.book["Summary"]
                    for cell in ws_sum[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor=header_fill_hex)
                        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                    ws_sum.freeze_panes = "A2"
                except Exception:
                    pass
                wrote_any_real_sheet = True
        except Exception:
            # Fallback placeholder failed; ignore to avoid aborting save
            pass
