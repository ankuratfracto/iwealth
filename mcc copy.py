from concurrent.futures import ThreadPoolExecutor, as_completed

# Process one page per chunk; still run up to 10 in parallel
CHUNK_SIZE_PAGES = 1
MAX_PARALLEL     = 10
MIN_TAIL_COMBINE = 1   # never merge tail pages so each page is sent individually

def _split_pdf_bytes(pdf_bytes: bytes,
                     chunk_size: int = CHUNK_SIZE_PAGES,
                     min_tail: int = MIN_TAIL_COMBINE) -> list[bytes]:
    """
    Return a list of PDF byte-chunks. Keeps 5-page blocks, *except* that a final
    fragment < min_tail pages is merged into the previous chunk so it retains
    invoice context (e.g. 26 pages → 5,5,5,5,6 instead of 5,5,5,5,5,1).
    """
    reader = PdfReader(io.BytesIO(pdf_bytes))
    total  = len(reader.pages)
    if total <= chunk_size:
        return [pdf_bytes]

    chunks: list[bytes] = []
    start = 0
    while start < total:
        end = min(start + chunk_size, total)
        # If this is the *last* chunk and it is tiny → back-merge with previous.
        if total - start < min_tail and chunks:
            # Append the remaining pages to the previous writer
            prev_buf = io.BytesIO(chunks.pop())          # last chunk bytes
            prev_reader = PdfReader(prev_buf)
            writer = PdfWriter()
            # re-copy pages of previous chunk
            for p in prev_reader.pages:
                writer.add_page(p)
            # add the tail pages
            for p in range(start, total):
                writer.add_page(reader.pages[p])
            buf = io.BytesIO()
            writer.write(buf)
            chunks.append(buf.getvalue())
            break   # finished
        else:
            writer = PdfWriter()
            for p in range(start, end):
                writer.add_page(reader.pages[p])
            buf = io.BytesIO()
            writer.write(buf)
            chunks.append(buf.getvalue())
            start = end

    return chunks

def call_fracto_parallel(pdf_bytes: bytes, file_name: str) -> list[dict]:
    """
    If the PDF is ≤ chunk_size_pages, behaves like `call_fracto` (returns [single‑result]).
    If more, splits into chunk_size_pages page chunks and hits the Fracto API concurrently with
    up to `MAX_PARALLEL` workers. Results are returned in order of the chunks.
    """
    chunks = _split_pdf_bytes(pdf_bytes, CHUNK_SIZE_PAGES)
    if len(chunks) == 1:
        return [call_fracto(pdf_bytes, file_name)]

    logger.info("Splitting %s into %d chunks of %d pages each", file_name, len(chunks), CHUNK_SIZE_PAGES)

    results: list[dict] = [None] * len(chunks)

    with ThreadPoolExecutor(max_workers=MAX_PARALLEL) as pool:
        futures = {
            pool.submit(call_fracto, chunk, f"{file_name} (part {i+1})"): i
            for i, chunk in enumerate(chunks)
        }
        for fut in as_completed(futures):
            idx = futures[fut]
            try:
                results[idx] = fut.result()
            except Exception as exc:
                logger.error("Chunk %d failed: %s", idx + 1, exc)
                results[idx] = {"file": file_name, "status": "error", "error": str(exc)}

    _renumber_serials(results)
    return results
#!/usr/bin/env python
"""
fracto_page_ocr.py
──────────────────
Split a PDF page-by-page and pipe each page through Fracto Smart-OCR.
"""

import io
import os
import sys
import json
import time
import logging
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import yaml

import requests
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas

# ─── PDF Stamping Helper ──────────────────────────────────────────────
def stamp_job_number(src_bytes: bytes, job_no: str, margin: int = 20) -> bytes:
    """
    Return new PDF bytes with an extra *margin* (pt) added to the top
    of every page, then stamps 'Job Number: <job_no>' inside that space.

    This ensures the stamp never covers the original page content.
    """
    if not job_no:
        return src_bytes

    from PyPDF2 import PdfReader, PdfWriter, Transformation, PageObject

    base_reader = PdfReader(io.BytesIO(src_bytes))
    writer      = PdfWriter()

    for orig_page in base_reader.pages:
        w = float(orig_page.mediabox.width)
        h = float(orig_page.mediabox.height)

        # 1️⃣  Create a new blank page taller by *margin*
        new_page = PageObject.create_blank_page(None, w, h + margin)

        # 2️⃣  Shift original page content down by `margin`
        orig_page.add_transformation(Transformation().translate(tx=0, ty=-margin))
        new_page.merge_page(orig_page)

        # 3️⃣  Create text overlay the same enlarged size
        overlay_buf = io.BytesIO()
        c = canvas.Canvas(overlay_buf, pagesize=(w, h + margin))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, h + margin - 15, f"Job Number: {job_no}")
        c.save()
        overlay_buf.seek(0)

        overlay_reader = PdfReader(overlay_buf)
        new_page.merge_page(overlay_reader.pages[0])

        # 4️⃣  Add to writer
        writer.add_page(new_page)

    out_buf = io.BytesIO()
    writer.write(out_buf)
    return out_buf.getvalue()

# ─── CONFIG ──────────────────────────────────────────────────────────────
FRACTO_ENDPOINT = "https://prod-ml.fracto.tech/upload-file-smart-ocr"
API_KEY         = os.getenv("FRACTO_API_KEY", "5d262446-4cfa-408b-86ef-04ad876bf2fc")
PARSER_APP_ID   = "dHbxlm0iggBuFgEZ"
MODEL_ID        = "gv1"
EXTRA_ACCURACY  = "true"

# ──────────────────────────────────────────────────────────────────────────

logger = logging.getLogger("FractoPageOCR")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)


def _load_formats():
    """
    Parse mapping.yaml and return a dict[str, dict] keyed by human‑friendly
    format name → {'mappings':…, 'template_path':…, 'sheet_name':…}
    and gracefully support three YAML layouts:
      ① legacy `excel_export` (single format)
      ② multiple `excel_export_*` siblings
      ③ modern `formats: { … }`
    """
    script_dir   = Path(__file__).parent
    mapping_file = script_dir / "mapping.yaml"
    formats: dict[str, dict] = {}

    if not mapping_file.exists():
        return formats

    with open(mapping_file, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}

    # ③ modern block
    if isinstance(data, dict) and "formats" in data:
        for name, cfg in data["formats"].items():
            if isinstance(cfg, dict):
                formats[str(name)] = cfg

    # ① legacy – keep as "Customs Invoice"
    if isinstance(data, dict) and "excel_export" in data:
        formats["Customs Invoice"] = data["excel_export"]

    # ② multiple excel_export_* blocks
    for key, val in (data.items() if isinstance(data, dict) else []):
        if key.startswith("excel_export_") and isinstance(val, dict):
            pretty = key.replace("excel_export_", "").replace("_", " ").title()
            formats[pretty] = val

    # Fallback: raw mapping dict alone
    if not formats and isinstance(data, dict):
        formats["Customs Invoice"] = {"mappings": data}

    # Normalise paths & ensure each entry has mappings dict
    for cfg in formats.values():
        if "mappings" not in cfg:
            cfg["mappings"] = {}
        if tp := cfg.get("template_path"):
            cfg["template_path"] = (script_dir / tp).expanduser().resolve()

    return formats

FORMATS = _load_formats()
DEFAULT_FORMAT = next(iter(FORMATS)) if FORMATS else "Customs Invoice"

# Keep legacy single‑format globals for existing callers
_default_cfg       = FORMATS.get(DEFAULT_FORMAT, {})
MAPPINGS           = _default_cfg.get("mappings", {})
TEMPLATE_PATH      = _default_cfg.get("template_path")
SHEET_NAME         = _default_cfg.get("sheet_name")
HEADERS = list(MAPPINGS.keys())


def call_fracto(file_bytes: bytes, file_name: str) -> Dict[str, Any]:
    """
    Send the whole PDF to Fracto OCR and return the JSON response.
    """
    files = {
        "file": (file_name, io.BytesIO(file_bytes), "application/pdf"),
    }
    data = {
        "parserApp": PARSER_APP_ID,
        "model": MODEL_ID,
        "extra_accuracy": EXTRA_ACCURACY,
    }
    headers = {"x-api-key": API_KEY}

    try:
        start = time.time()
        resp = requests.post(
            FRACTO_ENDPOINT,
            headers=headers,
            files=files,
            data=data,
            timeout=600,         # seconds
        )
        resp.raise_for_status()
        elapsed = time.time() - start
        logger.info("✓ %s processed in %.2fs", file_name, elapsed)
        return {"file": file_name, "status": "ok", "data": resp.json()}
    except Exception as exc:
        logger.error("✗ %s failed: %s", file_name, exc)
        return {"file": file_name, "status": "error", "error": str(exc)}




# ─── Helper to persist results ───────────────────────────────────────────
def save_results(results: List[Dict[str, Any]], pdf_path: str, out_path: str | None = None) -> str:
    """
    Persist OCR results to disk.

    If *out_path* is None, a file named "<original‑stem>_ocr.json" is created
    alongside the input PDF.

    Returns the absolute path to the saved file.
    """
    if out_path is None:
        p = Path(pdf_path).expanduser().resolve()
        out_path = p.with_name(f"{p.stem}_ocr.json")
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(results, fh, indent=2)
    logger.info("Results written to %s", out_path)
    return str(out_path)


# ─── Simple helper for CLI workflow ──────────────────────────────────────
def process_pdf(pdf_path: str) -> list[dict]:
    """
    Read *pdf_path* from disk and OCR it via `call_fracto_parallel`, honouring
    the current CHUNK_SIZE_PAGES and MAX_PARALLEL settings.

    Returns the list of Fracto API responses for every page‑chunk.
    """
    pdf_path = Path(pdf_path).expanduser().resolve()
    with open(pdf_path, "rb") as fh:
        pdf_bytes = fh.read()
    return call_fracto_parallel(pdf_bytes, pdf_path.name)

# ─── CLI ─────────────────────────────────────────────────────────────────
def _cli():
    """
    Usage:
        python -m mcc <pdf-path> [output.json] [output.xlsx] [KEY=VALUE ...]

    Convenience:
        • If you pass only two arguments and the second one ends with .xlsx / .xlsm / .xls,
          it is treated as the Excel output, and the JSON will default to
          "<pdf‑stem>_ocr.json" next to the PDF.
        • Any KEY=VALUE pairs will be written or overwritten in every row of the Excel output.
    """
    if len(sys.argv) < 2:
        print("Usage: python -m mcc <pdf-path> [output.json] [output.xlsx] [KEY=VALUE ...]")
        sys.exit(1)

    args = sys.argv[1:]

    pdf_path     = args[0]
    json_out     = None
    excel_out    = None

    # Collect KEY=VALUE overrides (e.g. --set Client=Acme or Client=Acme)
    overrides = {}
    remaining = []
    for arg in args[1:]:
        if "=" in arg:
            k, v = arg.split("=", 1)
            overrides[k.strip()] = v
        else:
            remaining.append(arg)
    # Re‑interpret remaining (non‑override) args for json/excel outputs
    if remaining:
        if remaining[0].lower().endswith((".xlsx", ".xlsm", ".xls")):
            excel_out = remaining[0]
        else:
            json_out = remaining[0]
    if len(remaining) >= 2:
        excel_out = remaining[1]

    if not os.path.isfile(pdf_path):
        logger.error("File not found: %s", pdf_path)
        sys.exit(2)

    results = process_pdf(pdf_path)

    # save JSON (use default if not supplied)
    save_results(results, pdf_path, json_out)

    # save Excel if requested
    if excel_out:
        write_excel_from_ocr(results, excel_out, overrides)




def _extract_rows(payload: Any) -> List[Dict[str, Any]]:
    """
    Heuristically extract a list[dict] rows from various JSON shapes
    that Fracto may return.

    • If *payload* is already a list of dicts → return as‑is.
    • If it's a dict, look for common keys ('data', 'rows', 'items', 'result', 'results')
      that hold a list of rows.
    • If the dict itself looks like a single row (has ≥1 HEADERS key) → wrap in list.
    Otherwise → empty list.
    """
    if isinstance(payload, list):
        return [r for r in payload if isinstance(r, dict)]

    if isinstance(payload, dict):
        for key in ("data", "rows", "items", "result", "results"):
            maybe = payload.get(key)
            if isinstance(maybe, list):
                return [r for r in maybe if isinstance(r, dict)]
        # treat dict itself as a single row if it shares keys
        if any(k in payload for k in HEADERS):
            return [payload]

    # Fallback: look inside `parsedData` for the first list of dicts
    if isinstance(payload, dict) and "parsedData" in payload:
        pd = payload["parsedData"]
        if isinstance(pd, dict):
            for v in pd.values():
                if isinstance(v, list) and v and isinstance(v[0], dict):
                    return [r for r in v if isinstance(r, dict)]

    return []  # fallback


def write_excel_from_ocr(
    results: List[Dict[str, Any]],
    output_path: str | io.BytesIO,
    overrides: dict[str, str] | None = None,
    *,
    mappings: dict[str, str] | None = None,
    template_path: str | None = None,
    sheet_name: str | None = None,
):
    """
    Write OCR rows to *output_path*.

    Parameters
    ----------
    results : list[dict]
        Fracto API responses (or pre‑loaded JSON) – list of results.
    output_path : str | io.BytesIO
        Where to write the Excel workbook.
    overrides : dict[str, str], optional
        {column_name: value} pairs forced into every row (e.g. constant HS‑Code).
    mappings : dict[str, str], optional
        Column → source‑field mapping. Defaults to the global MAPPINGS.
    template_path : str | Path, optional
        Path to an .xlsx template to use as a base (preserves styles).
    sheet_name : str, optional
        Which sheet inside the template/workbook to write into. Defaults to the
        first/active sheet.
    """
    mappings = mappings or MAPPINGS
    template_path = template_path or TEMPLATE_PATH
    sheet_name = sheet_name or SHEET_NAME
    overrides = overrides or {}

    headers = list(mappings.keys())

    # Keep only overrides whose column exists in the header list
    overrides = {k: v for k, v in overrides.items() if k in headers}

    # Load or create workbook
    if template_path and Path(template_path).expanduser().exists():
        wb = load_workbook(Path(template_path).expanduser())
    else:
        wb = Workbook()

    # Select or create target sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Clear existing data
    ws.delete_rows(1, ws.max_row)

    # ── Gather rows from results ──
    all_rows: list[dict] = []
    for result in results:
        payload = result.get("data", [])
        rows = _extract_rows(payload)
        all_rows.extend(rows)

    # Header row
    ws.append(headers)

    # Write data rows
    written = 0
    for row in all_rows:
        excel_row = []
        for col in headers:
            src_field = mappings.get(col, col)
            value = overrides.get(col, row.get(src_field, ""))
            excel_row.append(value)
        ws.append(excel_row)
        written += 1

    # ── Styling (same as before) ──
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="305496")
    header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
    thin_border  = Side(border_style="thin", color="999999")
    border       = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    max_width = 60
    for column in ws.iter_cols(min_row=1, max_row=ws.max_row):
        longest = max(len(str(c.value)) if c.value is not None else 0 for c in column)
        width = min(max(longest + 2, 10), max_width)
        ws.column_dimensions[column[0].column_letter].width = width
        for c in column[1:]:
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    # Zebra striping for readability
    stripe_fill = PatternFill("solid", fgColor="F2F2F2")
    for r in range(2, ws.max_row + 1):
        if r % 2 == 0:
            for c in ws[r]:
                c.fill = stripe_fill

    # Save
    if isinstance(output_path, io.BytesIO):
        wb.save(output_path)
    else:
        wb.save(str(output_path))

    logger.info(
        "Excel written to %s (%d rows, %d columns)",
        output_path if isinstance(output_path, str) else "<buffer>",
        written,
        len(headers),
    )

def _renumber_serials(results: list[dict],
                      json_field: str = "Serial_Number",
                      excel_header: str = "Item No.") -> None:
    """
    Mutates *results* in-place so that every row has a globally increasing
    serial number (1, 2, 3 …) across all Fracto chunks.

    The column name in the JSON is *json_field*; if it differs between your two
    formats, you can look it up via mappings in the caller instead.
    """
    counter = 1
    for res in results:
        rows = _extract_rows(res.get("data", []))
        for row in rows:
            row[json_field] = counter
            counter += 1


# ─── Main Entry Point ────────────────────────────────────────────────────
if __name__ == "__main__":
    _cli()