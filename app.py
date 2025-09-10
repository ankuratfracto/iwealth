"""Streamlit app for PDF → OCR → Excel with progress.

Interactive frontend that runs the iWealth OCR pipeline using the
`iwe_core` package, shows progress, and emits a styled statements
workbook plus a combined JSON. Intended for manual runs and demos.
"""

from __future__ import annotations
import logging
import re

# Use iwe_core directly (remove dependency on local a.py)
from iwe_core.config import CFG
from iwe_core.ocr_client import call_fracto_parallel, call_fracto
from iwe_core.excel_ops import (
    write_excel_from_ocr,
    generate_statements_excel,
    sanitize_statement_df,
    expand_selected_pages,
    _write_statements_workbook,
)
from iwe_core.grouping import (
    normalize_doc_type,
    extract_page_texts_from_pdf_bytes,
    infer_doc_type_from_text,
    build_groups,
)
from iwe_core.utils import (
    company_type_from_token,
    format_ranges,
    is_true_flag,
)
from iwe_core.json_ops import extract_rows as _extract_rows

from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

# ---- Helpers migrated from local a.py (config-driven formats, routing, company type) ----
import os
import io as _io
import pandas as _pd


def _load_formats_from_yaml() -> dict:
    """Load Excel output formats from mapping.yaml configured under CFG.paths.mapping_yaml."""
    mapping_rel = (CFG.get("paths", {}) or {}).get("mapping_yaml", "mapping.yaml")
    mapping_file = (Path(__file__).parent / mapping_rel).expanduser().resolve()
    formats: dict[str, dict] = {}
    if not mapping_file.exists():
        return formats
    try:
        import yaml as _yaml
        data = _yaml.safe_load(mapping_file.read_text(encoding="utf-8")) or {}
    except Exception:
        return formats

    if isinstance(data, dict) and "formats" in data:
        for name, cfg in (data.get("formats") or {}).items():
            if isinstance(cfg, dict):
                formats[str(name)] = cfg

    if isinstance(data, dict) and "excel_export" in data:
        formats["Customs Invoice"] = data["excel_export"]

    for key, val in (data.items() if isinstance(data, dict) else []):
        if isinstance(key, str) and key.startswith("excel_export_") and isinstance(val, dict):
            pretty = key.replace("excel_export_", "").replace("_", " ").title()
            formats[pretty] = val

    if not formats and isinstance(data, dict):
        formats["Customs Invoice"] = {"mappings": data}

    for cfg in formats.values():
        if "mappings" not in cfg:
            cfg["mappings"] = {}
        tp = cfg.get("template_path")
        if tp:
            cfg["template_path"] = (Path(__file__).parent / str(tp)).expanduser().resolve()
    return formats


FORMATS = _load_formats_from_yaml()


def normalize_company_type(ct_raw: str | None) -> str:
    s = re.sub(r"\s+", " ", (ct_raw or "").strip().lower())
    if not s:
        return str((CFG.get("company_type_prior", {}) or {}).get("default", "corporate")).lower()
    tokens = [t.strip() for t in re.split(r"[|/,;]+", s) if t.strip()]
    for tok in tokens:
        cls = company_type_from_token(tok)
        if cls:
            return cls
    if "bank" in s and "non banking" not in s and "non-banking" not in s:
        return "bank"
    if "nbfc" in s or "non banking financial" in s or "non-banking financial" in s:
        return "nbfc"
    if "insur" in s:
        return "insurance"
    if "non financial" in s or "corporate" in s or "company" in s:
        return "corporate"
    return str((CFG.get("company_type_prior", {}) or {}).get("default", "corporate")).lower()


# Routing config helpers
_ROUTING_CFG = CFG.get("routing", {}) or {}
_ROUTING_COMPANY_DEFAULT = str((CFG.get("company_type_prior", {}) or {}).get("default", "corporate")).lower()
_ROUTING_FALLBACK_ORDER = _ROUTING_CFG.get(
    "fallback_order",
    ["company_type_and_doc_type", "corporate_and_doc_type", "third_defaults"],
)
_ROUTING_ALLOWED_PARSERS = set(((_ROUTING_CFG.get("allowed_parsers") or []) or []))
_ROUTING_BLOCKED_PARSERS = set(((_ROUTING_CFG.get("blocked_parsers") or []) or []))
_ROUTING_SKIP_ON_DISABLED = bool(_ROUTING_CFG.get("skip_on_disabled", True))


def _resolve_routing(doc_type: str, company_type: str | None = None) -> tuple[str | None, str | None, str | None]:
    """Resolve parser/model/extra for a doc_type with detailed routing logs.

    Logs inputs, searched keys, allow/block effects, fallbacks attempted,
    and the final outcome. Uses INFO level to surface during normal runs.
    """
    logger = logging.getLogger(__name__)
    dt = (doc_type or "").strip().lower()
    ct = (company_type or _ROUTING_COMPANY_DEFAULT or "corporate").strip().lower()

    # Snapshot routing context for visibility
    try:
        _ct_keys = sorted(list((_ROUTING_CFG.get(ct) or {}).keys())) if isinstance(_ROUTING_CFG.get(ct), dict) else []
        _corp_keys = sorted(list((_ROUTING_CFG.get("corporate") or {}).keys())) if isinstance(_ROUTING_CFG.get("corporate"), dict) else []
        _allowed = sorted(list(_ROUTING_ALLOWED_PARSERS)) if _ROUTING_ALLOWED_PARSERS else []
        _blocked = sorted(list(_ROUTING_BLOCKED_PARSERS)) if _ROUTING_BLOCKED_PARSERS else []
        logger.info("[routing] inputs: company_type=%s doc_type=%s", ct, dt)
        logger.info("[routing] available doc_type keys → ct=%s: %s | corporate: %s", ct, _ct_keys, _corp_keys)
        if _allowed:
            logger.info("[routing] allowed_parsers=%s", _allowed)
        if _blocked:
            logger.info("[routing] blocked_parsers=%s", _blocked)
        logger.info("[routing] skip_on_disabled=%s; fallback_order=%s", _ROUTING_SKIP_ON_DISABLED, _ROUTING_FALLBACK_ORDER)
    except Exception:
        pass

    def _lookup(ct_key: str, dt_key: str):
        ct_map = _ROUTING_CFG.get(ct_key, {})
        if isinstance(ct_map, dict):
            hit = ct_map.get(dt_key)
            if isinstance(hit, dict):
                # Per-entry enable toggle
                if str(hit.get("enable", True)).strip().lower() in {"false", "0", "no", "off"}:
                    if _ROUTING_SKIP_ON_DISABLED:
                        try:
                            logger.info("[routing] %s/%s is disabled and skip_on_disabled=true → skipping", ct_key, dt_key)
                        except Exception:
                            pass
                        return (None, None, None)
                    try:
                        logger.info("[routing] %s/%s is disabled but skip_on_disabled=false → ignoring disable (will continue fallbacks)", ct_key, dt_key)
                    except Exception:
                        pass
                    return None
                parser = hit.get("parser") or (CFG.get("passes", {}).get("third", {}).get("defaults", {}) or {}).get("parser_app", "")
                model = hit.get("model") or (CFG.get("passes", {}).get("third", {}).get("defaults", {}) or {}).get("model", "tv6")
                extra = str(hit.get("extra", (CFG.get("passes", {}).get("third", {}).get("defaults", {}) or {}).get("extra_accuracy", True))).lower()
                # Global allow/block checks
                if _ROUTING_ALLOWED_PARSERS and parser not in _ROUTING_ALLOWED_PARSERS:
                    try:
                        logger.info("[routing] parser %s not in allowed_parsers; falling back", parser)
                    except Exception:
                        pass
                    return None
                if parser in _ROUTING_BLOCKED_PARSERS:
                    try:
                        logger.info("[routing] parser %s is blocked; falling back", parser)
                    except Exception:
                        pass
                    return None
                try:
                    logger.info("[routing] matched %s/%s → parser=%s, model=%s, extra=%s", ct_key, dt_key, parser, model, extra)
                except Exception:
                    pass
                return (parser, model, extra)
        return None

    for mode in _ROUTING_FALLBACK_ORDER:
        try:
            logger.info("[routing] attempt=%s ct=%s dt=%s", mode, ct, dt)
        except Exception:
            pass
        if mode == "company_type_and_doc_type":
            r = _lookup(ct, dt)
            if r == (None, None, None):
                return r
            if r:
                return r
        elif mode == "corporate_and_doc_type":
            r = _lookup("corporate", dt)
            if r:
                return r
        elif mode == "third_defaults":
            d = (CFG.get("passes", {}).get("third", {}).get("defaults", {}) or {})
            try:
                logger.info(
                    "[routing] no route; using third_defaults → parser=%s, model=%s, extra=%s",
                    d.get("parser_app", ""), d.get("model", "tv6"), str(d.get("extra_accuracy", True)).lower(),
                )
            except Exception:
                pass
            return (d.get("parser_app", ""), d.get("model", "tv6"), str(d.get("extra_accuracy", True)).lower())
    d = (CFG.get("passes", {}).get("third", {}).get("defaults", {}) or {})
    try:
        logger.info(
            "[routing] exhausted fallbacks; using third_defaults → parser=%s, model=%s, extra=%s",
            d.get("parser_app", ""), d.get("model", "tv6"), str(d.get("extra_accuracy", True)).lower(),
        )
    except Exception:
        pass
    return (d.get("parser_app", ""), d.get("model", "tv6"), str(d.get("extra_accuracy", True)).lower())

# --- Statements Excel with progress and concurrency ---
import re as _re

def reorder_dataframe_sections_first(df):
    """
    Ensure each section's header row appears before its break-up lines, with totals last.
    Heuristics:
      • A "header" row has a non-empty name column (Particulars/Description/etc.)
        and NO numeric values; it's not a Total/Subtotal.
      • "Total/Subtotal/Grand total" rows are pushed to the end of their section.
    """
    try:
        import pandas as _pd
    except Exception:
        return df
    if df is None or getattr(df, "empty", True):
        return df

    cols = list(df.columns)

    # 1) Find the "name" column
    name_col = None
    for c in cols:
        if str(c).strip().lower() in {"particulars","description","item","line_item","account","head","details"}:
            name_col = c
            break
    if name_col is None:
        return df  # no safe way to reorder

    # 2) Numeric columns (c1..cN or columns containing amount/value)
    num_cols = [c for c in cols if _re.fullmatch(r'(?i)c\d+', str(c)) or ("amount" in str(c).lower()) or ("value" in str(c).lower())]
    if not num_cols:
        meta = {name_col, "sr_no", "srno", "serial", "note"}
        num_cols = [c for c in cols if str(c).lower() not in {m.lower() for m in meta}]

    def _is_numlike(v):
        if v is None:
            return False
        if isinstance(v, str) and v.strip() in {"", "-", "–", "—", "na", "n/a", "nil"}:
            return False
        try:
            float(str(v).replace(",", ""))
            return True
        except Exception:
            return False

    n = len(df)
    is_header = [False]*n
    is_total  = [False]*n

    # optional sr_no helpers
    sr_cols = [c for c in cols if str(c).strip().lower() in {"sr_no","srno","serial","s no","s. no."}]
    sr_col = sr_cols[0] if sr_cols else None

    def cell(i, c):
        try:
            return df.iloc[i][c]
        except Exception:
            return None

    for i in range(n):
        name = str(cell(i, name_col) or "").strip()
        tot = bool(_re.match(r'^\s*(total|subtotal|grand\s+total)\b', name.lower()))
        is_total[i] = tot
        has_num = any(_is_numlike(cell(i, c)) for c in num_cols)

        hdr = (name != "") and (not has_num) and (not tot)
        if not hdr and sr_col:
            sr = str(cell(i, sr_col) or "").strip()
            if sr and _re.fullmatch(r'(?i)([ivxlcdm]+|\d+\.|[A-Za-z]\))', sr) and (not has_num):
                hdr = True
        is_header[i] = hdr

    out_idx, used = [], [False]*n

    def append_details(start, end):
        if start > end: return
        block = [i for i in range(start, end+1) if not is_header[i] and not used[i]]
        non_tot = [i for i in block if not is_total[i]]
        tots    = [i for i in block if is_total[i]]
        for i in non_tot + tots:
            out_idx.append(i); used[i] = True

    i = 0
    while i < n:
        if used[i]:
            i += 1; continue
        if is_header[i]:
            out_idx.append(i); used[i] = True
            j = i + 1
            while j < n and not is_header[j]:
                j += 1
            append_details(i+1, j-1)
            i = j
        else:
            # break-up before header → move next header before this block
            k = i
            while k < n and not is_header[k]:
                k += 1
            if k < n:
                out_idx.append(k); used[k] = True
                append_details(i, k-1)       # details that came before header
                j = k + 1
                while j < n and not is_header[j]:
                    j += 1
                append_details(k+1, j-1)     # details that follow header
                i = j
            else:
                append_details(i, n-1)
                i = n

    try:
        return df.iloc[out_idx].reset_index(drop=True)
    except Exception:
        return df

import os, time

def generate_statements_excel_with_progress(pdf_bytes: bytes, original_filename: str, progress, status_write):
    """Run the 1st/2nd/3rd pass with UI updates and concurrency; returns workbook bytes or None."""
    import pandas as pd  # local, avoids import-order issues
    import io            # local, avoids import-order issu
    t_overall = time.time()

    # 1) First pass
    status_write("1/3 First pass — per-page OCR …")
    t0 = time.time()
    results = call_fracto_parallel(
        pdf_bytes,
        original_filename,
        extra_accuracy=str(CFG.get("passes", {}).get("first", {}).get("extra_accuracy", False)).lower(),
    )
    dt0 = time.time() - t0
    progress.progress(0.33, text=f"First pass complete in {dt0:.1f}s")
    status_write(f"✓ First pass complete in {dt0:.1f}s — {len(results)} page(s)")

    # PDF page assembly handled by iwe_core.pdf_ops
    # 2) Select pages where has_table=true; also include pages whose headers clearly indicate a known doc type
    def _get_has_table(res: dict) -> bool:
        field = str(((CFG.get("passes", {}).get("first", {}).get("selection", {}) or {}).get("has_table_field", "has_table")))
        pdict = (res.get("data", {}) or {}).get("parsedData", {}) or {}
        if isinstance(pdict, list):
            for item in pdict:
                if isinstance(item, dict) and field in item:
                    return is_true_flag(item.get(field))
            return False
        return is_true_flag(pdict.get(field))
    selected_pages = [idx + 1 for idx, res in enumerate(results) if _get_has_table(res)]
    # Heuristic include: any page where header text suggests a known doc type (BS/PL/Cashflow)
    try:
        page_texts = extract_page_texts_from_pdf_bytes(pdf_bytes)
        header_pages = []
        for i, txt in enumerate(page_texts, start=1):
            inferred = infer_doc_type_from_text(txt)
            inferred_norm = normalize_doc_type(inferred) if inferred else None
            if inferred_norm and inferred_norm != "Others":
                header_pages.append(i)
        if header_pages:
            selected_pages = sorted(set(selected_pages) | set(header_pages))
    except Exception:
        pass
    if not selected_pages:
        status_write("⚠️ No pages flagged via table or headers — including all pages.")
        selected_pages = list(range(1, len(results) + 1))
    # Optional neighbour expansion via env var (default 0 to keep strict table-only selection)
    try:
        _radius = int(os.getenv("FRACTO_EXPAND_NEIGHBORS", str(int(((CFG.get("passes", {}).get("first", {}).get("selection", {}) or {}).get("neighbor_radius", 0))))) )
    except Exception:
        _radius = int(((CFG.get("passes", {}).get("first", {}).get("selection", {}) or {}).get("neighbor_radius", 0)))
    selected_pages = expand_selected_pages(selected_pages, len(results), radius=_radius)
    # Explicitly log which pages will go to the second pass
    try:
        logging.getLogger(__name__).info(
            "Second pass: re-processing %d selected pages %s",
            len(selected_pages), selected_pages,
        )
    except Exception:
        pass

    # Build selected.pdf
    from iwe_core.pdf_ops import build_pdf_from_pages
    selected_bytes = build_pdf_from_pages(pdf_bytes, selected_pages)

    # 3) Second pass
    status_write("2/3 Second pass — classifying selected pages …")
    stem = Path(original_filename).stem
    t1 = time.time()
    sel_pdf_name = str((CFG.get("passes", {}).get("second", {}) or {}).get("selected_pdf_name", "{stem}_selected.pdf")).format(stem=stem)
    second_res = call_fracto(
        selected_bytes,
        sel_pdf_name,
        parser_app=str((CFG.get("passes", {}).get("second", {}) or {}).get("parser_app", "")),
        model=str((CFG.get("passes", {}).get("second", {}) or {}).get("model", "tv7")),
        extra_accuracy=str((CFG.get("passes", {}).get("second", {}) or {}).get("extra_accuracy", False)).lower(),
    )
    dt1 = time.time() - t1
    progress.progress(0.55, text=f"Second pass complete in {dt1:.1f}s")
    status_write(f"✓ Second pass complete in {dt1:.1f}s")

    # 4) Classification (w/ fallback)
    pd_payload = (second_res.get("data", {}) or {}).get("parsedData", {})
    # Robust classification extraction (supports 'page_wise_classification' or 'classification')
    classification = []
    if isinstance(pd_payload, dict):
        classification = pd_payload.get("page_wise_classification") or pd_payload.get("classification") or []
    elif isinstance(pd_payload, list):
        classification = pd_payload
    else:
        classification = []
    norm_class = []
    for i, item in enumerate(classification, start=1):
        if not isinstance(item, dict):
            continue
        main_dt = item.get("doc_type") or item.get("Document_type")
            has_two = is_true_flag(item.get("has_two") or item.get("Has_multiple_sections"))
        second_dt = item.get("second_doc_type") or item.get("Second_doc_type")
        norm_class.append({
            "page_number": int(item.get("page_number") or i),
            "doc_type": main_dt,
            "has_two": "true" if has_two else "",
            "second_doc_type": second_dt,
            "is_continuation": "true" if is_true_flag(item.get("is_continuation")) else "",
            "continuation_of": item.get("continuation_of"),
        })
    classification = norm_class

    # Log the raw second-pass classification for visibility
    try:
        _dbg_cls = [
            {
                "page_number": int(it.get("page_number") or idx),
                "doc_type": it.get("doc_type") or it.get("Document_type"),
                "second_doc_type": it.get("second_doc_type") or it.get("Second_doc_type"),
            "has_two": is_true_flag(it.get("has_two") or it.get("Has_multiple_sections")),
            }
            for idx, it in enumerate(classification, start=1)
            if isinstance(it, dict)
        ]
        logging.getLogger(__name__).info("[routing] second-pass classification (raw) → %s", _dbg_cls)
    except Exception:
        pass

    # Company type for routing (from second pass, with normalisation)
    org_type_raw = None
    if isinstance(pd_payload, dict):
        try:
            # Prefer explicit 'organisation_type.type' if present
            org_type_raw = (pd_payload.get("organisation_type") or {}).get("type")
            if org_type_raw is not None:
                logging.getLogger(__name__).debug("[routing] organisation_type found via default path 'organisation_type.type': %r", org_type_raw)
            else:
                logging.getLogger(__name__).debug("[routing] organisation_type.type missing in payload; will use default company_type if needed")
        except Exception:
            pass
    company_type = normalize_company_type(org_type_raw)
    logging.getLogger(__name__).info("Routing company_type: %s (raw=%r)", company_type, org_type_raw)

    if not classification:
        classification = [
            {"page_number": i + 1, "doc_type": r.get("data", {}).get("parsedData", {}).get("Document_type")}
            for i, r in enumerate(results)
            if r.get("data", {}).get("parsedData", {}).get("Document_type", "Others").lower() != "others"
        ]
        classification = [it for it in classification if (it["page_number"] in selected_pages)]
    if not classification:
        status_write("⚠️ Could not derive classification — aborting third pass.")
        return None

    # Show mapping from selected-page index to original page number with labels
    try:
        _map = []
        for it in classification:
            sel_no = int((it or {}).get("page_number") or 0)
            if sel_no <= 0:
                continue
            if 1 <= sel_no <= len(selected_pages):
                orig_p = selected_pages[sel_no - 1]
            else:
                continue
            main = normalize_doc_type((it or {}).get("doc_type") or (it or {}).get("continuation_of"))
            sec  = normalize_doc_type((it or {}).get("second_doc_type") or "")
            flags = []
            if is_true_flag((it or {}).get("is_continuation")):
                flags.append("cont")
            if is_true_flag((it or {}).get("has_two")) or is_true_flag((it or {}).get("Has_multiple_sections")):
                flags.append("has_two")
            lab = main or "Others"
            if sec and sec != "Others" and sec != lab:
                lab = f"{lab} + {sec}"
            if flags:
                lab = f"{lab} ({','.join(flags)})"
            _map.append(f"{sel_no}→{orig_p}:{lab}")
        if _map:
            logging.getLogger(__name__).info("Second-pass mapping (sel→orig:label) → %s", "; ".join(_map))
    except Exception:
        pass

    # 5) Group pages by doc_type (robust: classification + header heuristics + smoothing)
    groups = build_groups(
        selected_pages, classification, pdf_bytes, first_pass_results=results
    )
    if not groups:
        status_write("⚠️ No groups found after classification.")
        return None

    # Log a one-line summary of all groups with page ranges
    try:
        _summary = "; ".join(
            f"{dt}: {format_ranges(sorted(pages))}" for dt, pages in sorted(groups.items())
        )
        logging.getLogger(__name__).info("[groups] Third pass groups → %s", _summary)
    except Exception:
        pass

    n_groups = len(groups)
    status_write(f"3/3 Third pass — {n_groups} document type(s): {sorted(groups.keys())}")

    # 6) Process groups concurrently (limit = MAX_PARALLEL)
    combined_sheets: dict[str, "pd.DataFrame"] = {}
    routing_used: dict[str, dict] = {}
    periods_hint: dict[str, dict] = {}
    completed = 0
    total = n_groups

    max_parallel = int((CFG.get("concurrency", {}) or {}).get("max_parallel", 9))
    with ThreadPoolExecutor(max_workers=min(max_parallel, n_groups)) as pool:
        futures = {}
        for doc_type, page_list in groups.items():
            page_list = sorted(page_list)
            from iwe_core.pdf_ops import build_pdf_from_pages
            group_bytes = build_pdf_from_pages(pdf_bytes, page_list)

            # Show the final routing keys we are about to use
            logging.getLogger(__name__).info(
                "[routing] inputs: ct_raw=%r ct=%s doc_raw=%r doc_norm=%s key=%s",
                org_type_raw, company_type, doc_type, doc_type, f"{company_type}/{doc_type.lower()}",
            )
            parser_app, model_id, extra_acc = _resolve_routing(doc_type, company_type=company_type)
            if parser_app is None:
                routing_used[doc_type] = {"parser_app": None, "model": None, "extra": None, "company_type": company_type, "skipped": True, "reason": "disabled"}
                logging.getLogger(__name__).info(
                    "↷ Skipping %s via company_type=%s (disabled; no fallback)",
                    doc_type, company_type
                )
                continue
            routing_used[doc_type] = {"parser_app": parser_app, "model": model_id, "extra": extra_acc, "company_type": company_type}
            logging.getLogger(__name__).info(
                "→ Routing %s via company_type=%s → parser=%s, model=%s, extra=%s, pages=%s",
                doc_type, company_type, parser_app, model_id, extra_acc, page_list
            )

            _slug = doc_type.lower().replace(' ', '_').replace('&','and').replace('/','_')
            _group_pdf_name = CFG.get("export", {}).get("filenames", {}).get("group_pdf", "{stem}_{slug}.pdf").format(stem=stem, slug=_slug)
            futures[pool.submit(
                call_fracto,
                group_bytes,
                _group_pdf_name,
                parser_app=parser_app,
                model=model_id,
                extra_accuracy=extra_acc,
            )] = doc_type

        for fut in as_completed(futures):
            doc_type = futures[fut]
            try:
                g0 = time.time()
                group_res = fut.result()
                gdt = time.time() - g0
                status_write(f"  ✓ {doc_type} done in {gdt:.1f}s")

                # Normalize group_res into a dict (API may return JSON string/bytes)
                if isinstance(group_res, (bytes, bytearray)):
                    try:
                        import json as _json
                        group_res = _json.loads(group_res.decode("utf-8", "ignore"))
                    except Exception:
                        group_res = {}
                elif isinstance(group_res, str):
                    try:
                        import json as _json
                        group_res = _json.loads(group_res)
                    except Exception:
                        group_res = {}

                if not isinstance(group_res, dict):
                    status_write(f"  ✗ {doc_type} returned non-JSON response; skipping.")
                    continue

                # Collect period metadata for this doc_type (id -> meta dict)
                try:
                    _periods = (((group_res or {}).get("data", {}) or {}).get("parsedData", {}) or {}).get("meta", {}).get("periods") or []
                    _by_id = {}
                    for _p in _periods:
                        if isinstance(_p, dict):
                            _pid = str((_p.get("id") or "")).strip().lower()
                            if _pid:
                                _by_id[_pid] = {
                                    "label": _p.get("label") or "",
                                    "start_date": _p.get("start_date"),
                                    "end_date": _p.get("end_date"),
                                    "role": _p.get("role"),
                                    "is_cumulative": _is_true(_p.get("is_cumulative")),
                                    "is_audited": _is_true(_p.get("is_audited")),
                                }
                    if _by_id:
                        periods_hint[doc_type] = _by_id
                except Exception:
                    pass

            except Exception as exc:
                status_write(f"  ✗ {doc_type} failed: {exc}")
                continue
            finally:
                completed += 1
                progress.progress(0.55 + 0.40 * (completed / total), text=f"Third pass {completed}/{total}: {doc_type}")

            parsed = group_res.get("data", {}).get("parsedData", [])
            rows_list = _extract_rows(parsed)
            if rows_list:
                all_keys = []
                for row in rows_list:
                    for k in row.keys():
                        if k not in all_keys:
                            all_keys.append(k)
                rows = [{k: r.get(k, "") for k in all_keys} for r in rows_list]
                df_ = pd.DataFrame(rows, columns=all_keys)
                df_ = sanitize_statement_df(doc_type, df_)
                # Preserve LLM order; only reorder when explicitly enabled
                try:
                    import os as _os
                    if str(_os.getenv("IWEALTH_ENABLE_REORDER", "0")).strip() in {"1", "true", "yes"}:
                        df_ = reorder_dataframe_sections_first(df_)
                except Exception:
                    pass
                combined_sheets[doc_type] = df_

    if not combined_sheets:
        status_write("⚠️ No tabular data parsed in third pass.")
        return None

    # 7) Write workbook using the shared CLI writer (single-pass, correct headers)
    import tempfile as _tempfile, shutil as _shutil, json as _json
    tmpdir = Path(_tempfile.mkdtemp(prefix="iwealth_")).resolve()
    try:
        # Save the uploaded PDF so the writer can resolve paths & infer BS labels if needed
        tmp_pdf_path = (tmpdir / f"{stem}.pdf")
        tmp_pdf_path.write_bytes(pdf_bytes)

        # Provide pages mapping so the writer can infer BS headers from PDF if periods are missing
        json_name_tmpl = CFG.get("export", {}).get("combined_json", {}).get("filename", "{stem}_statements.json")
        combined_obj = {"documents": {dt: {"pages": groups.get(dt, [])} for dt in groups}}
        (tmpdir / json_name_tmpl.format(stem=stem)).write_text(_json.dumps(combined_obj), encoding="utf-8")

        # Call the shared writer; it returns the xlsx path
        xlsx_out = _write_statements_workbook(
            str(tmp_pdf_path),
            stem,
            combined_sheets,
            routing_used=routing_used,
            periods_by_doc=periods_hint
        )
        xlsx_bytes = Path(xlsx_out).read_bytes()
        progress.progress(1.0, text=f"All done in {time.time()-t_overall:.1f}s")
        status_write("✅ Excel ready to download.")
        return xlsx_bytes
    finally:
        try:
            _shutil.rmtree(tmpdir)
        except Exception:
            pass

import io, textwrap
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import base64
import logging
# moved above to import from module `a`
from iwe_core.pdf_ops import get_page_count_from_stream

# ── Page config (must be first Streamlit command) ─────────────
st.set_page_config(
    page_title="PDF → Smart‑OCR → Excel",
    page_icon="📄",
    layout="wide",
)

# ── Fracto branding styles ────────────────────────────────────
FRACTO_PRIMARY   = "#00AB6B"   # adjust if brand palette differs
FRACTO_DARK      = "#00895A"
FRACTO_LIGHT_BG  = "#F5F8FF"

st.markdown(f"""
    <style>
    /* Page background */
    .stApp {{
        background: {FRACTO_LIGHT_BG};
    }}
    /* Center main content max-width 880px */
    .main .block-container{{
        max-width:880px;
        margin:auto;
    }}
    .block-container{{
        max-width:880px !important;
        margin-left:auto !important;
        margin-right:auto !important;
    }}
    /* Primary buttons */
    button[kind="primary"] {{
        background-color: {FRACTO_PRIMARY} !important;
        color: #fff !important;
        border: 0 !important;
    }}
    button[kind="primary"]:hover {{
        background-color: {FRACTO_DARK} !important;
        color: #fff !important;
    }}
    /* Header text color */
    h1 {{
        color: {FRACTO_DARK};
    }}
    /* Manual text_input boxes: white background & border */
    .stTextInput > div > div > input {{
        background-color: #ffffff !important;
        border: 1px solid #CCCCCC !important;
        border-radius: 4px !important;
    }}
    .stTextInput > div > div > input:focus {{
        border: 1px solid #00AB6B !important;   /* Fracto primary on focus */
        box-shadow: 0 0 0 2px rgba(0,171,107,0.2) !important;
    }}
    /* File uploader box */
    .stFileUploader > div > div {{
        background-color: #ffffff !important;
        border: 1px solid #CCCCCC !important;
        border-radius: 4px !important;
        color: #222222 !important;
    }}
    /* Fix inside text in uploader */
    .stFileUploader label {{
        color: #222222 !important;
    }}
    /* Force background and text for all blocks */
    html, body, .stApp, .block-container {{
        background-color: #FFFFFF !important;
        color: #222222 !important;
    }}
    /* Buttons in login section */
    button, .stButton button {{
        background-color: #00AB6B !important;
        color: #ffffff !important;
    }}
    button:hover, .stButton button:hover {{
        background-color: #00895A !important;
        color: #ffffff !important;
    }}
    /* Labels stay dark text */
    label, .stMarkdown, .stSubheader, .stHeader, .stTextInput label {{
        color: #222222 !important;
    }}
    /* Password input */
    input[type="password"] {{
        background-color: #FFFFFF !important;
        color: #222222 !important;
        border: 1px solid #CCCCCC !important;
    }}
    /* Dark mode: follow system preference */
    @media (prefers-color-scheme: dark) {{
        /* Base surfaces and text */
        html, body, .stApp, .block-container {{
            background-color: #0E1117 !important; /* Streamlit dark base */
            color: #E5E7EB !important;            /* slate-200 */
        }}
        .stApp {{
            background: #0E1117 !important;
        }}
        h1, h2, h3, h4, h5, h6 {{
            color: #F9FAFB !important;            /* near-white headers */
        }}
        p, span, div, label {{
            color: #E5E7EB !important;
        }}

        /* Inputs */
        .stTextInput > div > div > input,
        textarea {{
            background-color: #111827 !important; /* slate-900 */
            color: #E5E7EB !important;
            border: 1px solid #374151 !important; /* slate-700 */
        }}
        input[type="password"] {{
            background-color: #111827 !important;
            color: #E5E7EB !important;
            border: 1px solid #374151 !important;
        }}

        /* File uploader */
        .stFileUploader > div > div {{
            background-color: #111827 !important;
            border: 1px solid #374151 !important;
            color: #E5E7EB !important;
        }}
        .stFileUploader label {{
            color: #E5E7EB !important;
        }}

        /* Buttons (keep brand green) */
        button[kind="primary"], .stButton button {{
            background-color: #00AB6B !important;
            color: #FFFFFF !important;
            border: 0 !important;
        }}
        button[kind="primary"]:hover, .stButton button:hover {{
            background-color: #00895A !important;
        }}

        /* Cards and panels */
        .card {{
            background: #1F2937;                   /* slate-800 */
            border: 1px solid #374151;            /* slate-700 */
            box-shadow: 0 1px 3px rgba(0,0,0,0.35);
        }}
        .card h4 {{
            color: #34D399;                        /* teal-300 */
        }}
    }}
    /* Force hover shadow on cards */
    .card:hover {{
        box-shadow: 0 4px 8px rgba(0,0,0,0.15);
        transition: box-shadow 0.3s ease-in-out;
    }}
    /* Scrolling logo strip */
    .logo-strip-wrapper{{
        max-width:880px;
        margin:24px auto;
        overflow:hidden;
    }}
    .logo-strip{{
        display:inline-block;
        white-space:nowrap;
        animation:logoscroll 20s linear infinite; /* doubled speed */
    }}
    .logo-strip img{{
        height:48px;
        margin:0 32px;
        vertical-align:middle;
        display:inline-block;
    }}
    /* Remove extra gap where the duplicated sequence joins */
    .logo-strip img:last-child{{
        margin-right:0;
    }}
    /* Remove margin-left on first clone to shorten overall gap */
    .logo-strip img:nth-child(1){{
        margin-left:0;
    }}
    @keyframes logoscroll{{
        0%   {{transform:translateX(0);}}
        100% {{transform:translateX(-50%);}}
    }}
    </style>
""", unsafe_allow_html=True)
# ── Clients logo strip ───────────────────────────────────────
def build_logo_strip(logo_paths: list[str]) -> str:
    """
    Return HTML for the scrolling logo strip.
    Each file is read from disk and embedded as a Base64 data‑URI,
    so it renders correctly on Streamlit Cloud.
    """
    tags = ""
    script_dir = Path(__file__).parent
    for rel_path in logo_paths:
        img_path = (script_dir / rel_path).expanduser().resolve()
        if not img_path.exists():
            continue
        mime = "image/svg+xml" if img_path.suffix.lower() == ".svg" else "image/png"
        try:
            b64 = base64.b64encode(img_path.read_bytes()).decode("utf-8")
        except Exception as e:
            logging.warning("Failed to read logo %s: %s", img_path, e)
            continue
        tags += f"<img src='data:{mime};base64,{b64}' alt='' />"
    # Duplicate sequence so the CSS animation loops seamlessly
    return f"<div class='logo-strip-wrapper'><div class='logo-strip'>{tags}{tags}</div></div>"

st.markdown(
    """
    <style>
    .card-container {
        display: flex;
        gap: 1rem;
        flex-wrap: wrap;
        margin-bottom: 1rem;
    }
    .card {
        flex: 1 1 200px;
        background: #F6F8FA;
        border: 1px solid #E0E0E0;
        border-radius: 12px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        padding: 1rem;
        text-align: center;
    }
    .card-icon {
        margin-bottom: 8px;
        display:flex;
        justify-content:center;
    }
    .card-icon img{
        width:36px;
        height:36px;
    }
    .card h4{
        font-size:16px;
        font-weight:600;
        margin:4px 0 8px 0;
        color: #00895A;
    }
    .card p{
        font-size:13px;
        line-height:1.4rem;
        margin:0;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# Logo banner at the top
st.image("fractologo.jpeg", width=180)

# ── Session keys ─────────────────────────────────────────────
if "excel_bytes" not in st.session_state:
    st.session_state["excel_bytes"] = None
if "excel_filename" not in st.session_state:
    st.session_state["excel_filename"] = ""
if "edited_excel_bytes" not in st.session_state:
    st.session_state["edited_excel_bytes"] = None
    st.session_state["edited_filename"] = ""

# ── Simple username/password gate ─────────────────────────────
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    st.subheader("🔐 Login required")
    uname = st.text_input("Username")
    pword = st.text_input("Password", type="password")
    if st.button("Login"):
        if uname == "iwealth" and pword == "iwealth@99":
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Invalid credentials")
    st.stop()   # prevent the rest of the app from rendering

# Ensure FRACTO_API_KEY is available for API calls
if "FRACTO_API_KEY" in st.secrets:
    os.environ["FRACTO_API_KEY"] = st.secrets["FRACTO_API_KEY"]



# ── Hero / intro ──────────────────────────────────────────────
st.markdown(
    '''
    <h2 style="color:#00895A;font-weight:600;margin-bottom:0.2rem;">Automate imports. Eliminate re‑typing.</h2>
    <p style="font-size:1.05rem;line-height:1.5rem;margin-bottom:1.5rem;">
      Fracto converts your shipping invoices, customs docs and purchase orders into<br>
      ERP‑ready spreadsheets in seconds — complete with your business rules and validation checks.
    </p>
    ''',
    unsafe_allow_html=True,
)
# 24px spacing before uploader
st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

st.markdown("## Smart‑OCR to ERP‑ready Excel")

st.markdown('<h3 id="upload">1. Upload and process your PDF</h3>', unsafe_allow_html=True)

# Output mode toggle — default to Statements Excel (like CLI third pass)
use_statements_mode = st.checkbox(
    "Statements Excel (like CLI third pass)", value=True,
    help="Generates the same multi-sheet workbook grouped by document type. Uncheck to use a single-sheet format from mapping.yaml."
)

# ── Upload & Process ──────────────────────────────────────────
# Upload widget
pdf_file = st.file_uploader("Upload PDF", type=["pdf"])

# Show thumbnail info after upload
if pdf_file:
    # Show file thumbnail info
    file_size_kb = pdf_file.size / 1024
    try:
        page_count = get_page_count_from_stream(pdf_file)
    except Exception:
        page_count = "?"
    st.info(f"**{pdf_file.name}**  •  {file_size_kb:,.1f} KB  •  {page_count} page(s)")
    # Reset file pointer for later reading
    pdf_file.seek(0)

st.markdown("#### Optional manual fields")
# (Hidden for now) No manual UI fields; you can still set overrides programmatically if needed.
manual_inputs: dict[str, str] = {}

selected_format_cfg = None
if not use_statements_mode:
    # Formats come straight from mapping.yaml ("Format 1", "Format 2", …)
    format_names = list(FORMATS.keys())
    selected_format_key = st.selectbox("Select Excel output format", format_names)
    selected_format_cfg = FORMATS[selected_format_key]

# Process button
run = st.button("⚙️ Process PDF", disabled=pdf_file is None)

if run:
    if not pdf_file:
        st.warning("Please upload a PDF first.")
        st.stop()

    progress = st.progress(0.0, text="Uploading & extracting …")
    try:
        pdf_bytes = pdf_file.read()
        progress.progress(0.2)
        progress.progress(0.4)

        excel_bytes = None
        base_name = Path(pdf_file.name).stem

        if use_statements_mode:
            with st.status("Processing…", expanded=True) as status_box:
                excel_bytes = generate_statements_excel_with_progress(pdf_bytes, pdf_file.name, progress, status_box.write)

        if excel_bytes is None:
            # Fallback to single-sheet mapping export
            progress.progress(0.6, text="Extracting rows (single-sheet)…")
            results = call_fracto_parallel(pdf_bytes, pdf_file.name)
            progress.progress(0.8)

            buffer = io.BytesIO()
            # Pick a default format if none was selected (e.g., checkbox toggled mid-run)
            if selected_format_cfg is None:
                if FORMATS:
                    default_key = next(iter(FORMATS))
                    selected_format_cfg = FORMATS[default_key]
                else:
                    selected_format_cfg = {"mappings": {}, "template_path": None, "sheet_name": None}

            write_excel_from_ocr(
                results,
                buffer,
                overrides=manual_inputs,
                mappings=selected_format_cfg.get("mappings", {}),
                template_path=selected_format_cfg.get("template_path"),
                sheet_name=selected_format_cfg.get("sheet_name"),
            )
            excel_bytes = buffer.getvalue()
            final_name = f"{base_name}_ocr.xlsx"
        else:
            _tmpl = CFG.get("export", {}).get("filenames", {}).get("statements_xlsx", "{stem}_statements.xlsx")
            final_name = _tmpl.format(stem=base_name)

        progress.progress(1.0, text="Done!")
        st.session_state["excel_bytes"]   = excel_bytes
        st.session_state["excel_filename"] = final_name
        st.toast("✅ Excel generated!", icon="🎉")
    except Exception as exc:
        st.toast(f"❌ Error: {exc}", icon="⚠️")
        st.error(f"Processing failed: {exc}")
        st.stop()

# ── Preview & download ────────────────────────────────────────
if st.session_state["excel_bytes"]:
    st.markdown("### 2. Review and export")
    st.download_button(
        "⬇️ Download original Excel",
        data=st.session_state["excel_bytes"],
        file_name=st.session_state["excel_filename"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_original",
    )

    # Support multi-sheet workbooks (like Statements). Default to first non-summary sheet.
    try:
        xls = pd.ExcelFile(io.BytesIO(st.session_state["excel_bytes"]), engine="openpyxl")
    except Exception as e:
        st.error(f"Could not open the generated Excel: {e}")
        st.stop()
    all_sheets = xls.sheet_names
    editable_sheets = [s for s in all_sheets if s.lower() != "routing summary"] or all_sheets

    if "selected_sheet" not in st.session_state:
        st.session_state["selected_sheet"] = editable_sheets[0]

    st.write("**Select sheet to review/edit**")
    selected_sheet = st.selectbox("Sheet", editable_sheets, index=editable_sheets.index(st.session_state["selected_sheet"]))
    st.session_state["selected_sheet"] = selected_sheet

    df = pd.read_excel(xls, sheet_name=selected_sheet)
    edited_df = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        key=f"editable_grid_{selected_sheet}",
    )

    if st.button("💾 Save edits"):
        from openpyxl import load_workbook
        wb_orig = load_workbook(io.BytesIO(st.session_state["excel_bytes"]))
        ws      = wb_orig[selected_sheet]

        # Update header to match edited columns (keeps existing cell styles)
        for c_idx, col_name in enumerate(list(edited_df.columns), start=1):
            ws.cell(row=1, column=c_idx, value=str(col_name))

        # Overwrite data rows using fast itertuples (values only)
        for r_idx, row in enumerate(edited_df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Trim any leftover rows below the new data
        last_data_row = edited_df.shape[0] + 1  # header is row 1
        if ws.max_row > last_data_row:
            ws.delete_rows(last_data_row + 1, ws.max_row - last_data_row)

        out_buf = io.BytesIO()
        wb_orig.save(out_buf)
        st.session_state["edited_excel_bytes"] = out_buf.getvalue()
        st.session_state["edited_filename"] = (
            Path(st.session_state["excel_filename"]).with_suffix("").name + f"_{selected_sheet}_edited.xlsx"
        )
        st.success(f"Edits to '{selected_sheet}' saved — scroll below to download the .xlsx file.")

    if st.session_state.get("edited_excel_bytes"):
        st.download_button(
            "⬇️ Download edited Excel",
            data=st.session_state["edited_excel_bytes"],
            file_name=st.session_state["edited_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_edited",
        )

    # ── Quick stats & visualisations ───────────────────────────
    view_df = edited_df if st.session_state.get("edited_excel_bytes") else df

    st.markdown("### 3. Quick stats")
    k1, k2 = st.columns(2)
    k1.metric("Total rows", view_df.shape[0])
    k2.metric("Blank cells", int(view_df.isna().sum().sum()))

    # Optionally show numeric totals if columns exist
    if "Qty" in view_df.columns:
        st.metric("Total Qty", f"{view_df['Qty'].sum():,.0f}")
    if "Unit Price" in view_df.columns:
        total_unit_price = (
            pd.to_numeric(view_df["Unit Price"], errors="coerce").fillna(0).sum()
        )
        st.metric("Sum Unit Price", f"{total_unit_price:,.0f}")

    # ── Top Part Numbers by Qty chart ─────────────────────────
    if {"Part No.", "Qty"}.issubset(view_df.columns):
        st.markdown("#### Top SKUs by Qty")
        top_qty = (
            view_df.groupby("Part No.")["Qty"].sum(numeric_only=True).sort_values(ascending=False).head(10)
        )
        if top_qty.empty or top_qty.shape[0] < 1:
            st.info("No Qty data available to plot.")
        else:
            fig, ax = plt.subplots()
            top_qty.plot(kind="barh", ax=ax)
            ax.invert_yaxis()
            ax.set_xlabel("Qty")
            ax.set_ylabel("Part No.")
            st.pyplot(fig)

st.markdown("---")

# ── Clients logo strip ───────────────────────────────────────
st.markdown("### Our Clients")
logo_files = [
    "clients/kuhoo.png",
    "clients/ODeX.png",
    "clients/accomation.png",
    "clients/jaikisan.png",
    "clients/121Finance.png",
    "clients/NBHC.png",
    "clients/MCC.png",
    "clients/navata.png",
    "clients/trukker.png",
    "clients/turno.png",
    "clients/petpooja.png",
    "clients/freightfox.png",
    "clients/presolv.png",
    "clients/equal.png",
    "clients/ambit.png",
    "clients/khfl.png",
    "clients/pssc.png",
    "clients/symbo.png",
]
st.markdown(build_logo_strip(logo_files), unsafe_allow_html=True)
st.markdown("---")

# ── Benefits grid ─────────────────────────────────────────────
st.markdown("### Why choose **Fracto Imports**?")
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown("#### 🚀 10× Faster")
    st.write("Upload → processed Excel in under a minute, even for multi‑page PDFs.")
with col2:
    st.markdown("#### 🔍 Error‑free")
    st.write("AI‑assisted extraction + your manual overrides ensure 99.9 % accuracy.")
with col3:
    st.markdown("#### 🔗 Fits Your ERP")
    st.write("Column mapping matches your import template out‑of‑the‑box.")

st.markdown("---")


# ── Inline SVG icons (Tabler, 36×36, stroke‑currentColor) ─
ICONS = {
    "upload": '''
      <svg xmlns="http://www.w3.org/2000/svg" width="36" height="36" viewBox="0 0 24 24" stroke="#00895A" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M4 17v2a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2v-2" /><polyline points="7 9 12 4 17 9" /><line x1="12" y1="4" x2="12" y2="16" /></svg>
    ''',
    "cpu": '''
      <svg xmlns="http://www.w3.org/2000/svg" width="36" height="36" viewBox="0 0 24 24" stroke="#00895A" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="4" y="4" width="16" height="16" rx="1" /><rect x="9" y="9" width="6" height="6" rx="1" /><path d="M3 9h1" /><path d="M3 15h1" /><path d="M20 9h1" /><path d="M20 15h1" /><path d="M9 3v1" /><path d="M15 3v1" /><path d="M9 20v1" /><path d="M15 20v1" /></svg>
    ''',
    "edit": '''
      <svg xmlns="http://www.w3.org/2000/svg" width="36" height="36" viewBox="0 0 24 24" stroke="#00895A" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M16 3l4 4l-11 11h-4v-4z" /><path d="M13 6l4 4" /><path d="M3 20v1h1l3-3" /></svg>
    ''',
    "export": '''
      <svg xmlns="http://www.w3.org/2000/svg" width="36" height="36" viewBox="0 0 24 24" stroke="#00895A" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 3v4a1 1 0 0 0 1 1h4" /><path d="M5 12v-7a2 2 0 0 1 2 -2h7l5 5v4" /><path d="M9 15l3 -3l3 3" /><path d="M12 12v9" /></svg>
    ''',
    "ship": '''
      <svg xmlns="http://www.w3.org/2000/svg" width="36" height="36" viewBox="0 0 24 24" stroke="#00895A" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 9l9 -4l9 4l-9 4z" /><path d="M3 9l9 4l9 -4" /><path d="M12 19l0 -11" /><path d="M9 21l-1 -7" /><path d="M15 21l1 -7" /></svg>
    ''',
    "factory": '''
      <svg xmlns="http://www.w3.org/2000/svg" width="36" height="36" viewBox="0 0 24 24" stroke="#00895A" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 21v-13l8 -4v7l8 -4v14" /><path d="M13 13l-8 -4" /><path d="M5 17h2v4h-2z" /><path d="M9 17h2v4h-2z" /><path d="M13 17h2v4h-2z" /><path d="M17 17h2v4h-2z" /></svg>
    ''',
    "dollar": '''
      <svg xmlns="http://www.w3.org/2000/svg" width="36" height="36" viewBox="0 0 24 24" stroke="#00895A" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 3v18" /><path d="M17 8a5 5 0 0 0 -10 0c0 5 5 3 10 8a5 5 0 0 1 -10 0" /></svg>
    ''',
}

# ── Card rendering helper ─────────────────────────────────────
def render_card(icon_name: str, title: str, body: str, *, width="250px") -> str:
    svg = ICONS.get(icon_name, "")
    return f"""
        <div class="card" style="max-width:{width};">
          <div class="card-icon">{svg}</div>
          <h4>{title}</h4>
          <p>{body}</p>
        </div>
    """

# ── How it works ──────────────────────────────────────────────
st.markdown('<h3 id="how">How it works</h3>', unsafe_allow_html=True)

steps = [
    ("upload", "Upload", "Drag PDFs or images of invoices, POs, customs docs into the drop‑zone."),
    ("cpu", "AI Extraction", "Reads tables, handwriting and stamps with 99 %+ accuracy."),
    ("edit", "Review & Edit", "Adjust any field inline — spreadsheet‑style editor keeps you in control."),
    ("export", "Export", "Download ERP‑ready Excel or push straight into your system via API."),
]

cols = st.columns(4)
for col, (icon_name, title, body) in zip(cols, steps):
    with col:
        col.markdown(render_card(icon_name, title, body), unsafe_allow_html=True)

st.markdown("---")

# ── Popular use‑cases ─────────────────────────────────────────
st.markdown('<h3 id="usecases">Popular use‑cases</h3>', unsafe_allow_html=True)

use_cases = [
    ("ship", "Import Logistics", "Bills of lading, packing lists, HS‑code mapping — ready for customs clearance."),
    ("factory", "Manufacturing", "Supplier invoices and QC sheets flow directly into SAP/Oracle with serial‑level traceability."),
    ("dollar", "Finance & AP", "Reconcile bank statements and purchase invoices 10× faster with zero manual key‑in."),
]

uc_cols = st.columns(3)
for col, (icon_name, title, body) in zip(uc_cols, use_cases):
    with col:
        col.markdown(render_card(icon_name, title, body, width="280px"), unsafe_allow_html=True)

st.markdown("---")

# ── Footer ────────────────────────────────────────────────────
st.markdown(
    "<div style='text-align:center;font-size:0.85rem;padding-top:2rem;color:#666;'>"
    "Made with ❤️ by <a href='https://www.fracto.tech' style='color:#00AB6B;' target='_blank'>Fracto</a>"
    "</div>",
    unsafe_allow_html=True,
)
